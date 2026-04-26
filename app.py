from datetime import datetime, date, timezone, timedelta
import base64
import csv
import io
import os
import re
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import altair as alt
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# ── Zona horaria Colombia (UTC-5, sin horario de verano) ─────────────────────
COL_TZ = timezone(timedelta(hours=-5))

def now_col() -> datetime:
    """Hora actual en zona horaria Colombia."""
    return datetime.now(tz=COL_TZ)

# ── Persistencia CSV ─────────────────────────────────────────────────────────
CSV_PATH  = "rutas_historial.csv"
FOTOS_DIR = "fotos"
os.makedirs(FOTOS_DIR, exist_ok=True)
CSV_COLS = [
    "tipo_seguimiento",
    "fecha", "ruta", "placa", "conductor",
    "volumen_declarado", "vol_estaciones", "diferencia",
    "solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
    "num_estaciones", "guardado_en",
    "st_carrotanque", "grasa_muestra", "proteina_muestra", "diferencia_solidos",
    "estaciones_json", "fotos_json",
]

# ── CSV separado para SEGUIMIENTOS ───────────────────────────────────────────
SEG_CSV_PATH = "seguimientos_historial.csv"
SEG_COLS = [
    "sub_tipo_seguimiento", "fecha",
    "seg_codigo", "seg_quien_trajo", "ruta", "seg_responsable",
    "seg_id_muestra", "seg_volumen", "seg_grasa", "seg_st", "seg_ic", "seg_agua",
    "seg_alcohol", "seg_cloruros", "seg_neutralizantes", "seg_observaciones",
    "seg_vol_declarado", "seg_vol_muestras", "seg_diferencia_vol",
    "seg_solidos_ruta", "seg_crioscopia_ruta", "seg_st_pond", "seg_ic_pond",
    "muestras_json", "guardado_en", "fotos_json",
]

DRAFT_PATH      = "borrador_autoguardado.json"
CATALOGO_PATH   = "estaciones_catalogo.csv"
DRAFT_EXACT_KEYS = [
    "continuar", "_tipo_servicio_guardado", "_sub_tipo_seg_guardado",
    "tipo_servicio_select", "sub_tipo_seg_select",
    "_ruta_fg",
    "imagenes_confirmadas", "imagenes_nombres_guardados",
    "trans_imagenes_confirmadas", "trans_imagenes_nombres_guardados",
    "estaciones_guardadas", "form_ver",
    "trans_fecha", "trans_placa", "trans_st_carrotanque",
    "trans_grasa", "trans_st_muestra", "trans_proteina",
    "seg_fecha", "seg_codigo", "seg_quien_trajo", "seg_ruta_acomp",
    "seg_responsable", "seg_quality_key_counter",
    "acomp_muestras", "contra_muestras",
]
DRAFT_PREFIXES = (
    "nue_",
    "fecha_ruta_", "nombre_ruta_", "placa_vehiculo_", "conductor_",
    "volumen_ruta_", "solidos_totales_", "crioscopia_",
    "seg_id_muestra_", "seg_grasa_", "seg_st_", "seg_ic_raw_", "seg_agua_",
    "seg_alcohol_", "seg_cloruros_", "seg_neutralizantes_", "seg_observaciones_",
)


def _draft_encode(value):
    if isinstance(value, datetime):
        return {"__draft_type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"__draft_type": "date", "value": value.isoformat()}
    try:
        json.dumps(value)
        return value
    except TypeError:
        return str(value)


def _draft_decode(value):
    if isinstance(value, dict) and value.get("__draft_type") in ("date", "datetime"):
        raw = value.get("value", "")
        try:
            return datetime.fromisoformat(raw).date()
        except Exception:
            return date.today()
    return value


def restore_draft_state():
    if st.session_state.get("_draft_restored"):
        return
    st.session_state["_draft_restored"] = True
    if not os.path.exists(DRAFT_PATH):
        return
    try:
        with open(DRAFT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return
    if "tipo_servicio_select" not in data and "_tipo_servicio_guardado" in data:
        data["tipo_servicio_select"] = data.get("_tipo_servicio_guardado")
    if "sub_tipo_seg_select" not in data and "_sub_tipo_seg_guardado" in data:
        data["sub_tipo_seg_select"] = data.get("_sub_tipo_seg_guardado")
    for key, value in data.items():
        if key not in st.session_state:
            st.session_state[key] = _draft_decode(value)


def save_draft_state():
    if st.session_state.pop("_skip_draft_save_once", False):
        return
    data = {}
    for key in DRAFT_EXACT_KEYS:
        if key in st.session_state:
            data[key] = _draft_encode(st.session_state[key])
    for key, value in st.session_state.items():
        if key.startswith(DRAFT_PREFIXES):
            data[key] = _draft_encode(value)
    try:
        with open(DRAFT_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def clear_draft_state():
    st.session_state["_skip_draft_save_once"] = True
    try:
        if os.path.exists(DRAFT_PATH):
            os.remove(DRAFT_PATH)
    except Exception:
        pass


def save_fotos_to_disk(uploaded_files: list, prefix: str) -> list[str]:
    """Guarda imágenes en FOTOS_DIR y retorna la lista de rutas relativas."""
    saved = []
    if not uploaded_files:
        return saved
    ts = now_col().strftime("%Y%m%d_%H%M%S")
    safe_prefix = re.sub(r"[^A-Z0-9_\-]", "_", prefix.upper())
    for i, uf in enumerate(uploaded_files, start=1):
        ext = uf.name.rsplit(".", 1)[-1].lower()
        ext = ext if ext in ("jpg", "jpeg", "png") else "jpg"
        fname = f"{safe_prefix}_{ts}_{i}.{ext}"
        fpath = os.path.join(FOTOS_DIR, fname)
        uf.seek(0)
        with open(fpath, "wb") as fh:
            fh.write(uf.read())
        uf.seek(0)
        saved.append(fpath)
    return saved


def load_historial() -> pd.DataFrame:
    if not os.path.exists(CSV_PATH):
        return pd.DataFrame(columns=CSV_COLS)
    try:
        df = pd.read_csv(CSV_PATH, dtype=str)
        # Columnas nuevas que pueden no existir en CSVs anteriores
        for col in CSV_COLS:
            if col not in df.columns:
                df[col] = ""
        # Tipo de seguimiento: filas vacías → RUTAS (compatibilidad)
        df["tipo_seguimiento"] = df["tipo_seguimiento"].fillna("RUTAS").replace("", "RUTAS")
        for col in ["volumen_declarado", "vol_estaciones", "diferencia", "num_estaciones"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        for col in ["solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
                    "st_carrotanque", "grasa_muestra", "proteina_muestra", "diferencia_solidos"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        if "fecha" in df.columns:
            df["_fecha_dt"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y", errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=CSV_COLS)


def save_ruta_to_csv(row: dict):
    if os.path.exists(CSV_PATH):
        df = load_historial()
        if "_fecha_dt" in df.columns:
            df = df.drop(columns=["_fecha_dt"])
    else:
        df = pd.DataFrame(columns=CSV_COLS)
    # Asegurar que todas las columnas existen en df
    for col in CSV_COLS:
        if col not in df.columns:
            df[col] = ""
    new_row = pd.DataFrame([{k: row.get(k, "") for k in CSV_COLS}])
    df = pd.concat([df[CSV_COLS], new_row], ignore_index=True)
    df.to_csv(CSV_PATH, index=False, encoding="utf-8")


def load_seguimientos() -> pd.DataFrame:
    if not os.path.exists(SEG_CSV_PATH):
        return pd.DataFrame(columns=SEG_COLS)
    try:
        df = pd.read_csv(SEG_CSV_PATH, dtype=str)
        for col in SEG_COLS:
            if col not in df.columns:
                df[col] = ""
        for col in ["seg_grasa", "seg_st", "seg_ic", "seg_agua"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        # Unificación: registros viejos guardados como TERCEROS → ESTACIONES
        if "sub_tipo_seguimiento" in df.columns:
            df["sub_tipo_seguimiento"] = df["sub_tipo_seguimiento"].replace("TERCEROS", "ESTACIONES")
        if "fecha" in df.columns:
            df["_fecha_dt"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y", errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=SEG_COLS)


@st.cache_data(ttl=300)
def load_catalogo() -> pd.DataFrame:
    """Carga el catálogo de estaciones. Retorna DataFrame vacío si no existe."""
    if not os.path.exists(CATALOGO_PATH):
        return pd.DataFrame(columns=["codigo", "nombre"])
    try:
        df = pd.read_csv(CATALOGO_PATH, dtype=str)
        df["codigo"] = df["codigo"].str.strip()
        df["nombre"] = df["nombre"].str.strip().str.upper()
        return df.dropna(subset=["codigo", "nombre"])
    except Exception:
        return pd.DataFrame(columns=["codigo", "nombre"])


def save_seguimiento_to_csv(row: dict):
    if os.path.exists(SEG_CSV_PATH):
        df = load_seguimientos()
        if "_fecha_dt" in df.columns:
            df = df.drop(columns=["_fecha_dt"])
    else:
        df = pd.DataFrame(columns=SEG_COLS)
    for col in SEG_COLS:
        if col not in df.columns:
            df[col] = ""
    new_row = pd.DataFrame([{k: row.get(k, "") for k in SEG_COLS}])
    df = pd.concat([df[SEG_COLS], new_row], ignore_index=True)
    df.to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")


def delete_seg_row(orig_idx: int):
    df = load_seguimientos()
    df = df.drop(index=orig_idx)
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[SEG_COLS].to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")


def delete_seg_rows(orig_indices: list):
    df = load_seguimientos()
    df = df.drop(index=[i for i in orig_indices if i in df.index])
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[SEG_COLS].to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")




def delete_row_from_csv(orig_idx: int):
    df = load_historial()
    df = df.drop(index=orig_idx)
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def delete_rows_from_csv(orig_indices: list):
    df = load_historial()
    df = df.drop(index=[i for i in orig_indices if i in df.index])
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def update_row_in_csv(orig_idx: int, new_vals: dict):
    df = load_historial()
    for k, v in new_vals.items():
        if k in df.columns:
            df.at[orig_idx, k] = v
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def calcular_estado_calidad(row: dict) -> str:
    """Retorna 'CONFORME' o 'DESVIACIÓN' según los parámetros de la ruta.
    Solo aplica a registros de tipo RUTAS."""
    if str(row.get("tipo_seguimiento", "RUTAS")).strip() != "RUTAS":
        return "CONFORME"
    try:
        st_val = float(str(row.get("solidos_ruta", "")).replace(",", "."))
        if 0 < st_val < 12.60:
            return "DESVIACIÓN"
    except (ValueError, TypeError):
        pass
    try:
        ic_val = float(str(row.get("crioscopia_ruta", "")).replace(",", "."))
        if ic_val > -0.535 or ic_val < -0.550:
            return "DESVIACIÓN"
    except (ValueError, TypeError):
        pass
    return "CONFORME"


def historial_to_excel_filtrado(
    df_filtrado: pd.DataFrame,
    fecha_desde,
    fecha_hasta,
    filtro_tipo: str,
) -> bytes:
    """Excel multi-hoja respetando los filtros del Historial."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_hdr = PatternFill("solid", fgColor="1F4E79")
    fill_bad = PatternFill("solid", fgColor="FFC7CE")
    fill_alt = PatternFill("solid", fgColor="EEF4FB")
    font_hdr = Font(bold=True, size=10, color="FFFFFF")
    font_bad = Font(bold=True, size=10, color="9C0006")
    bold     = Font(bold=True, size=10)
    normal   = Font(size=10)
    center   = Alignment(horizontal="center", vertical="center")
    bd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _wh(ws, cols, widths):
        for ci, hdr in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=hdr)
            c.fill = fill_hdr; c.font = font_hdr
            c.alignment = center; c.border = bd
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20

    def _wc(ws, ri, ci, val, fmt=None, bad=False, alt=False):
        v = val if (val is not None and not (isinstance(val, float) and pd.isna(val))) else ""
        c = ws.cell(row=ri, column=ci, value=v)
        c.alignment = center; c.border = bd
        if bad:
            c.font = font_bad; c.fill = fill_bad
        else:
            c.font = normal
            if alt: c.fill = fill_alt
        if fmt: c.number_format = fmt

    # ── Hoja RUTAS ────────────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "RUTAS"):
        df_r = (df_filtrado[df_filtrado["tipo_seguimiento"] == "RUTAS"].copy()
                if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws1 = wb.create_sheet("RUTAS")
        cols1 = [
            ("TIPO", "tipo_seguimiento"), ("FECHA", "fecha"), ("RUTA", "ruta"),
            ("PLACA", "placa"), ("CONDUCTOR", "conductor"),
            ("VOL. DECLARADO (L)", "volumen_declarado"),
            ("VOL. ESTACIONES (L)", "vol_estaciones"),
            ("DIFERENCIA (L)", "diferencia"),
            ("SÓLIDOS RUTA (%)", "solidos_ruta"),
            ("CRIOSCOPIA RUTA (°C)", "crioscopia_ruta"),
            ("ST POND", "st_pond"), ("IC POND", "ic_pond"),
            ("Nº ESTACIONES", "num_estaciones"), ("GUARDADO EN", "guardado_en"),
        ]
        _wh(ws1, [h for h, _ in cols1],
            [10, 12, 18, 10, 18, 16, 18, 14, 16, 18, 10, 10, 12, 18])
        for ri, row in enumerate(df_r.itertuples(index=False), start=2):
            rd = row._asdict()
            desv_st = desv_ic = False
            try:
                v = float(str(rd.get("solidos_ruta","")).replace(",","."))
                if 0 < v < 12.60: desv_st = True
            except Exception: pass
            try:
                v = float(str(rd.get("crioscopia_ruta","")).replace(",","."))
                if v > -0.535 or v < -0.550: desv_ic = True
            except Exception: pass
            alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols1, 1):
                fmt = "0.00"  if col in ("solidos_ruta","st_pond") else \
                      "0.000" if col in ("crioscopia_ruta","ic_pond") else None
                bad = ((desv_st or desv_ic) and col == "ruta") or \
                      (desv_st and col == "solidos_ruta") or \
                      (desv_ic and col == "crioscopia_ruta")
                _wc(ws1, ri, ci, rd.get(col,""), fmt=fmt, bad=bad, alt=alt)

    # ── Hoja TRANSUIZA ────────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "TRANSUIZA"):
        df_t = (df_filtrado[df_filtrado["tipo_seguimiento"] == "TRANSUIZA"].copy()
                if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws2 = wb.create_sheet("TRANSUIZA")
        cols2 = [
            ("FECHA","fecha"), ("PLACA","placa"),
            ("ST CARROTANQUE (%)","st_carrotanque"),
            ("GRASA (%)","grasa_muestra"), ("ST MUESTRA (%)","solidos_ruta"),
            ("PROTEÍNA (%)","proteina_muestra"),
            ("DIFERENCIA SÓLIDOS","diferencia_solidos"), ("GUARDADO EN","guardado_en"),
        ]
        _wh(ws2, [h for h, _ in cols2], [12, 10, 18, 10, 16, 12, 18, 18])
        for ri, row in enumerate(df_t.itertuples(index=False), start=2):
            rd = row._asdict(); alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols2, 1):
                fmt = "0.00" if col in ("st_carrotanque","grasa_muestra","solidos_ruta",
                                        "proteina_muestra","diferencia_solidos") else None
                _wc(ws2, ri, ci, rd.get(col,""), fmt=fmt, alt=alt)

    # ── Hoja SEGUIMIENTOS ─────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "SEGUIMIENTOS"):
        if filtro_tipo == "SEGUIMIENTOS":
            df_seg = df_filtrado.copy()
        else:
            df_seg = load_seguimientos()
            if "_fecha_dt" in df_seg.columns:
                df_seg = df_seg[
                    (df_seg["_fecha_dt"].dt.date >= fecha_desde) &
                    (df_seg["_fecha_dt"].dt.date <= fecha_hasta)
                ]
        df_seg = df_seg.drop(columns=["_fecha_dt","_estado"], errors="ignore")
        ws3 = wb.create_sheet("SEGUIMIENTOS")
        cols3 = [
            ("SUB-TIPO","sub_tipo_seguimiento"), ("FECHA","fecha"),
            ("CÓDIGO","seg_codigo"), ("ENTREGADO POR","seg_quien_trajo"),
            ("RUTA","ruta"), ("RESPONSABLE","seg_responsable"),
            ("ID MUESTRA","seg_id_muestra"), ("GRASA (%)","seg_grasa"),
            ("ST (%)","seg_st"), ("IC (°C)","seg_ic"), ("AGUA (%)","seg_agua"),
            ("ALCOHOL","seg_alcohol"), ("CLORUROS","seg_cloruros"),
            ("NEUTRALIZANTES","seg_neutralizantes"),
            ("OBSERVACIONES","seg_observaciones"), ("GUARDADO EN","guardado_en"),
        ]
        _wh(ws3, [h for h, _ in cols3],
            [18, 12, 12, 18, 16, 18, 14, 10, 10, 10, 10, 12, 12, 16, 30, 18])
        for ri, row in enumerate(df_seg.itertuples(index=False), start=2):
            rd = row._asdict(); alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols3, 1):
                fmt = "0.00"  if col in ("seg_grasa","seg_st","seg_agua") else \
                      "0.000" if col == "seg_ic" else None
                _wc(ws3, ri, ci, rd.get(col,""), fmt=fmt, alt=alt)

    # ── Hoja ESTACIONES ───────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "RUTAS"):
        df_re = (df_filtrado[df_filtrado["tipo_seguimiento"] == "RUTAS"].copy()
                 if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws4 = wb.create_sheet("ESTACIONES")
        hdrs4 = [
            "FECHA","RUTA","PLACA","CONDUCTOR","VOL. DECLARADO",
            "# ESTACIÓN","CÓDIGO","GRASA (%)","SÓL.TOT. (%)","PROTEÍNA (%)",
            "CRIOSCOPIA (°C)","VOLUMEN (L)","ALCOHOL","CLORUROS","NEUTRALIZANTES",
            "% AGUA","OBSERVACIONES","ST RUTA (%)","IC RUTA (°C)","ESTADO CALIDAD",
        ]
        _wh(ws4, hdrs4,
            [12,18,10,18,14,10,14,10,10,10,14,12,10,10,14,8,26,12,12,14])
        est_ri = 2
        for _, ruta_row in df_re.iterrows():
            raw_json = str(ruta_row.get("estaciones_json","") or "")
            try: ests = json.loads(raw_json) if raw_json.strip() else []
            except Exception: ests = []
            if not ests: continue
            try: st_rv = float(str(ruta_row.get("solidos_ruta","")).replace(",","."))
            except Exception: st_rv = None
            try: ic_rv = float(str(ruta_row.get("crioscopia_ruta","")).replace(",","."))
            except Exception: ic_rv = None
            desv_st_r = st_rv is not None and 0 < st_rv < 12.60
            desv_ic_r = ic_rv is not None and (ic_rv > -0.535 or ic_rv < -0.550)
            estado_r = "DESVIACIÓN" if (desv_st_r or desv_ic_r) else "CONFORME"
            for idx_e, est in enumerate(ests, 1):
                try: ic_e = float(str(est.get("crioscopia","")).replace(",",".")); desv_ic_e = ic_e > -0.535 or ic_e < -0.550
                except Exception: ic_e = None; desv_ic_e = False
                try: st_e = float(str(est.get("solidos","")).replace(",",".")); desv_st_e = 0 < st_e < 12.60
                except Exception: st_e = None; desv_st_e = False
                hay_desv_e = desv_ic_e or desv_st_e
                row_vals = [
                    ruta_row.get("fecha",""), ruta_row.get("ruta",""),
                    ruta_row.get("placa",""), ruta_row.get("conductor",""),
                    ruta_row.get("volumen_declarado",""), idx_e,
                    est.get("codigo",""), est.get("grasa"), est.get("solidos"),
                    est.get("proteina"), est.get("crioscopia"), est.get("volumen"),
                    est.get("alcohol",""), est.get("cloruros",""), est.get("neutralizantes",""),
                    est.get("agua_pct"), est.get("obs",""), st_rv, ic_rv, estado_r,
                ]
                fmts = [None,None,None,None,"0","0",None,
                        "0.00","0.00","0.00","0.000","0",
                        None,None,None,"0.0",None,"0.00","0.000",None]
                alt = (est_ri % 2 == 0)
                for ci_e, (val_e, fmt_e) in enumerate(zip(row_vals, fmts), 1):
                    bad_e = hay_desv_e and ci_e in (8, 9, 11)
                    _wc(ws4, est_ri, ci_e, val_e, fmt=fmt_e, bad=bad_e, alt=alt)
                est_ri += 1

    # ── Hoja ACOMPAÑAMIENTOS (muestras expandidas con ponderados) ─────────
    if filtro_tipo in ("TODOS", "SEGUIMIENTOS"):
        if filtro_tipo == "SEGUIMIENTOS":
            df_acomp_xl = df_filtrado.copy()
        else:
            df_acomp_xl = load_seguimientos()
            if "_fecha_dt" in df_acomp_xl.columns:
                df_acomp_xl = df_acomp_xl[
                    (df_acomp_xl["_fecha_dt"].dt.date >= fecha_desde) &
                    (df_acomp_xl["_fecha_dt"].dt.date <= fecha_hasta)
                ]
        df_acomp_xl = df_acomp_xl[
            df_acomp_xl.get("sub_tipo_seguimiento", pd.Series(dtype=str)) == "ACOMPAÑAMIENTOS"
        ] if "sub_tipo_seguimiento" in df_acomp_xl.columns else pd.DataFrame()
        if not df_acomp_xl.empty:
            try:
                _cat_xl = load_catalogo()
                _cat_xl_map = dict(zip(_cat_xl["codigo"], _cat_xl["nombre"]))
            except Exception:
                _cat_xl_map = {}
            ws5 = wb.create_sheet("ACOMPAÑAMIENTOS")
            hdrs5 = [
                "FECHA","RUTA","ENTREGADO POR","RESPONSABLE",
                "VOL. DECLARADO (L)","VOL. SUMA MUESTRAS (L)","DIFERENCIA (L)",
                "ST RUTA (%)","IC RUTA (°C)","ST PONDERADO (%)","IC PONDERADO (°C)",
                "# MUESTRA","CÓDIGO","NOMBRE ESTACIÓN","VOLUMEN (L)",
                "GRASA (%)","ST (%)","PROTEÍNA (%)","IC (°C)","AGUA (%)","POND ST","IC POND",
                "ALCOHOL","CLORUROS","NEUTRALIZANTES","OBSERVACIONES","GUARDADO EN",
            ]
            widths5 = [12,18,18,18,16,18,14,12,14,14,14,10,14,20,10,10,10,10,10,10,10,10,10,10,14,30,18]
            _wh(ws5, hdrs5, widths5)
            _am_ri = 2
            for _, _arow in df_acomp_xl.iterrows():
                _raw_mj = str(_arow.get("muestras_json","") or "")
                try: _muestras_xl = json.loads(_raw_mj) if _raw_mj.strip() else []
                except Exception: _muestras_xl = []
                def _pnxl(x):
                    try: return float(str(x).replace(",","."))
                    except: return None
                try: _st_rv = float(str(_arow.get("seg_solidos_ruta","")).replace(",","."))
                except: _st_rv = None
                try: _ic_rv = float(str(_arow.get("seg_crioscopia_ruta","")).replace(",","."))
                except: _ic_rv = None
                try: _st_pv = float(str(_arow.get("seg_st_pond","")).replace(",","."))
                except: _st_pv = None
                try: _ic_pv = float(str(_arow.get("seg_ic_pond","")).replace(",","."))
                except: _ic_pv = None
                try: _vol_decl_xl = int(float(str(_arow.get("seg_vol_declarado","")).replace(",",".")))
                except: _vol_decl_xl = None
                try: _vol_sum_xl  = int(float(str(_arow.get("seg_vol_muestras","")).replace(",",".")))
                except: _vol_sum_xl = None
                try: _dif_xl = int(float(str(_arow.get("seg_diferencia_vol","")).replace(",",".")))
                except: _dif_xl = None
                _common5 = [
                    _arow.get("fecha",""), _arow.get("ruta",""),
                    _arow.get("seg_quien_trajo",""), _arow.get("seg_responsable",""),
                    _vol_decl_xl, _vol_sum_xl, _dif_xl,
                    _st_rv, _ic_rv, _st_pv, _ic_pv,
                ]
                _alt5 = (_am_ri % 2 == 0)
                if not _muestras_xl:
                    for _ci5, _v5 in enumerate(_common5 + ["—"]*14 + [_arow.get("guardado_en","")], 1):
                        _wc(ws5, _am_ri, _ci5, _v5, alt=_alt5)
                    _am_ri += 1
                    continue
                for _idx5, _am5 in enumerate(_muestras_xl, 1):
                    _cod5 = str(_am5.get("ID","") or "").strip()
                    _vol5 = _pnxl(_am5.get("_volumen"))
                    _st5  = _pnxl(_am5.get("_st"))
                    _ic5  = _pnxl(_am5.get("_ic"))
                    _pst5 = round(_vol5 * _st5, 2) if _vol5 is not None and _st5 is not None else None
                    _pic5 = round(_vol5 * _ic5, 3) if _vol5 is not None and _ic5 is not None else None
                    _row5 = _common5 + [
                        _idx5, _cod5, _cat_xl_map.get(_cod5,""),
                        int(_vol5) if _vol5 is not None else None,
                        _pnxl(_am5.get("_grasa")), _st5, _pnxl(_am5.get("_proteina")), _ic5,
                        _pnxl(_am5.get("_agua")), _pst5, _pic5,
                        _am5.get("_alcohol",""), _am5.get("_cloruros",""),
                        _am5.get("_neutralizantes",""), _am5.get("_obs",""),
                        _arow.get("guardado_en",""),
                    ]
                    _alt5 = (_am_ri % 2 == 0)
                    for _ci5, _v5 in enumerate(_row5, 1):
                        _hdr5 = hdrs5[_ci5-1]
                        _fmt5 = "0.00"  if _hdr5 in ("ST RUTA (%)","ST PONDERADO (%)","GRASA (%)","ST (%)","PROTEÍNA (%)","AGUA (%)","POND ST") else \
                                "0.000" if _hdr5 in ("IC RUTA (°C)","IC PONDERADO (°C)","IC (°C)","IC POND") else None
                        _wc(ws5, _am_ri, _ci5, _v5, fmt=_fmt5, alt=_alt5)
                    _am_ri += 1

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



st.set_page_config(page_title="QualiLact", page_icon="🧪", layout="wide")
restore_draft_state()

st.markdown(
    """
    <style>
    /* ── Indicador de carga: ciclo de iconos ───────────────────── */
    @keyframes _ql_icono {
        0%,  30%  { content: "🐄  Cargando..."; }
        33%, 63%  { content: "🥛  Cargando..."; }
        66%, 96%  { content: "🧪  Cargando..."; }
        100%      { content: "🐄  Cargando..."; }
    }
    /* Ocultar spinner original */
    [data-testid="stStatusWidget"] > div > div { display: none !important; }
    /* Mostrar icono ciclico fijo en parte superior derecha */
    [data-testid="stStatusWidget"]::after {
        content: "🐄  Cargando...";
        position: fixed;
        top: 10px;
        right: 24px;
        font-size: 1rem;
        font-weight: 600;
        color: #0056A3;
        background: #EAF1FA;
        border: 1.5px solid #BDD7EE;
        border-radius: 20px;
        padding: 4px 14px;
        display: block;
        animation: _ql_icono 2.4s linear infinite;
        line-height: 1.6;
        z-index: 9999;
        pointer-events: none;
        white-space: nowrap;
    }


    /* ── Fondo azul claro QualiLact ────────────────────────────── */
    .stApp { background-color: #F0F5F9 !important; }
    section[data-testid="stSidebar"] { background-color: #E8EFF5 !important; }
    /* ── Tarjetas blancas redondeadas ──────────────────────────── */
    div[data-testid="stVerticalBlock"] > div[data-testid="element-container"],
    div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #FFFFFF;
    }
    section.main > div { background-color: #F0F5F9 !important; }

    /* ── Ocultar spinners de number input ──────────────────────── */
    input[type="number"]::-webkit-outer-spin-button,
    input[type="number"]::-webkit-inner-spin-button {
        -webkit-appearance: none; margin: 0;
    }
    input[type="number"] { -moz-appearance: textfield; }
    button[data-testid="stNumberInputStepUp"],
    button[data-testid="stNumberInputStepDown"] { display: none !important; }

    /* ── Inputs: fondo gris claro + bordes redondeados ─────────── */
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stDateInput"] input {
        background-color: #F4F4F4 !important;
        border: 1.5px solid #D1D5DB !important;
        border-radius: 8px !important;
        color: #1F2937 !important;
        font-size: 14px !important;
        padding: 8px 12px !important;
        transition: border-color 0.18s, box-shadow 0.18s;
    }

    /* ── Focus: resaltado azul Nestlé ──────────────────────────── */
    div[data-testid="stTextInput"] input:focus,
    div[data-testid="stNumberInput"] input:focus,
    div[data-testid="stDateInput"] input:focus {
        border-color: #0056A3 !important;
        box-shadow: 0 0 0 3px rgba(0, 86, 163, 0.12) !important;
        background-color: #FFFFFF !important;
        outline: none !important;
    }

    /* ── Labels: gris oscuro, peso semibold ────────────────────── */
    div[data-testid="stTextInput"] label p,
    div[data-testid="stNumberInput"] label p,
    div[data-testid="stDateInput"] label p,
    div[data-testid="stSelectbox"] label p {
        color: #555555 !important;
        font-weight: 600 !important;
        font-size: 12.5px !important;
        letter-spacing: 0.3px;
    }

    /* ── Selectbox: mismos bordes redondeados ──────────────────── */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:first-child {
        background-color: #F4F4F4 !important;
        border: 1.5px solid #D1D5DB !important;
        border-radius: 8px !important;
    }

    /* ── Divisores ─────────────────────────────────────────────── */
    hr { border-color: #E5E7EB !important; }

    /* ── Botones primarios: azul Nestlé ────────────────────────── */
    button[kind="primary"], button[data-testid*="primary"] {
        background-color: #0056A3 !important;
        border-color: #0056A3 !important;
        border-radius: 8px !important;
    }
    button[kind="primary"]:hover {
        background-color: #004285 !important;
    }

    /* ── Contenedores con borde ────────────────────────────────── */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 10px !important;
        border-color: #E5E7EB !important;
        background-color: #FAFAFA !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div style="
        display: flex;
        align-items: center;
        gap: 16px;
        padding: 18px 0 10px 0;
        border-bottom: 2px solid #0056A3;
        margin-bottom: 18px;
    ">
        <div style="
            font-size: 2.6rem;
            line-height: 1;
            letter-spacing: -2px;
        ">🐄🥛</div>
        <div>
            <div style="
                font-size: 2rem;
                font-weight: 800;
                color: #0056A3;
                letter-spacing: 1px;
                line-height: 1.1;
                font-family: 'Segoe UI', sans-serif;
            ">QualiLact</div>
            <div style="
                font-size: 0.9rem;
                color: #6B7280;
                font-weight: 400;
                letter-spacing: 0.5px;
                margin-top: 2px;
                font-family: 'Segoe UI', sans-serif;
            ">Control de Calidad en Leche Fresca</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if "continuar" not in st.session_state:
    st.session_state.continuar = False

if "pagina_activa" not in st.session_state:
    st.session_state.pagina_activa = "REGISTRAR"

for _sk, _sv in [
    ("admin_accion", None), ("admin_idx", None), ("admin_idxs", []),
    ("hist_buscar_ok", False),
    ("tipo_registrar", "RUTAS"), ("sub_tipo_registrar", "ESTACIONES"),
    ("registrar_submenu_open", False),
]:
    if _sk not in st.session_state:
        st.session_state[_sk] = _sv


def convertir_a_mayusculas(campo):
    st.session_state[campo] = st.session_state[campo].upper()


def sanitizar_nombre_ruta(campo):
    val = st.session_state.get(campo, "")
    st.session_state[campo] = re.sub(r"[^A-ZÁÉÍÓÚÑÜ0-9]", "", val.upper())


def validar_placa():
    _k = f"placa_vehiculo_{st.session_state.get('_ruta_fg', 0)}"
    if _k in st.session_state:
        st.session_state[_k] = re.sub(
            r"[^A-Z0-9]", "", st.session_state[_k].upper()
        )


def activar_siguiente_con_enter():
    components.html(
        """
        <script>
        function obtenerInputsVisibles() {
            return Array.from(window.parent.document.querySelectorAll("input, textarea"))
                .filter(el => {
                    const tipo = el.getAttribute("type");
                    if (tipo === "hidden" || tipo === "checkbox" || tipo === "radio") return false;
                    const rect = el.getBoundingClientRect();
                    return rect.width > 0 && rect.height > 0;
                });
        }

        function clickBtnGuardar(input) {
            // Primero busca AGREGAR MUESTRA (ACOMPAÑAMIENTOS); si no, GUARDAR
            const botones = Array.from(window.parent.document.querySelectorAll("button"));
            const visible  = b => b.offsetParent !== null && b.innerText;
            const btn = botones.find(b => visible(b) && b.innerText.includes("AGREGAR MUESTRA"))
                     || botones.find(b => visible(b) && b.innerText.includes("GUARDAR"));
            if (btn) setTimeout(() => { btn.click(); }, 80);
        }

        function moverFoco(input, delta) {
            const actualizados = obtenerInputsVisibles();
            const posActual = actualizados.indexOf(input);
            if (posActual === -1) return;          // elemento ya no está en el DOM
            const destino = actualizados[posActual + delta];
            if (destino) {
                setTimeout(() => {
                    destino.focus();
                    if (destino.select) destino.select();
                }, 60);
            }
        }

        function activarNavegacion() {
            const inputs = obtenerInputsVisibles();
            inputs.forEach((input) => {
                if (input.dataset.navActivo === "true") return;
                input.dataset.navActivo = "true";

                const esNumerico = input.getAttribute("type") === "number";

                input.addEventListener("keydown", (e) => {
                    // Enter → siguiente campo;
                    // si es OBSERVACIONES o es el último campo visible, pulsa GUARDAR
                    if (e.key === "Enter") {
                        e.preventDefault();
                        const ph = (input.placeholder || "").toLowerCase();
                        const todos = obtenerInputsVisibles();
                        const posActual = todos.indexOf(input);
                        const esUltimo = posActual !== -1 && posActual === todos.length - 1;
                        if (ph.includes("observaciones") || ph.includes("ingrese observaciones") || esUltimo) {
                            clickBtnGuardar(input);
                        } else {
                            moverFoco(input, 1);
                        }
                        return;
                    }
                    // ←→: en numéricos navegan siempre; en texto solo en los extremos
                    if (e.key === "ArrowRight") {
                        if (esNumerico || input.selectionStart >= input.value.length) {
                            e.preventDefault();
                            moverFoco(input, 1);
                        }
                    } else if (e.key === "ArrowLeft") {
                        if (esNumerico || input.selectionStart <= 0) {
                            e.preventDefault();
                            moverFoco(input, -1);
                        }
                    }
                    // ↑↓ sin función de navegación (comportamiento nativo del navegador)
                });
            });
        }

        activarNavegacion();
        setInterval(() => { activarNavegacion(); }, 600);
        </script>
        """,
        height=0,
    )

# ── SIDEBAR DE NAVEGACIÓN ────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        """
        <div style="text-align:center;padding:12px 0 18px 0;
                    border-bottom:2px solid #0056A3;margin-bottom:22px;">
            <div style="font-size:2rem;line-height:1.1;">🐄🥛</div>
            <div style="font-size:1.3rem;font-weight:800;color:#0056A3;
                        font-family:'Segoe UI',sans-serif;letter-spacing:1px;">
                QualiLact
            </div>
            <div style="font-size:0.72rem;color:#6B7280;margin-top:2px;">
                Control de Calidad Láctea
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <style>
        /* Sidebar nav buttons — full width, styled */
        [data-testid="stSidebar"] .stButton > button {
            width: 100%;
            border-radius: 8px;
            font-weight: 600;
            font-size: 0.95rem;
            padding: 10px 14px;
            margin-bottom: 6px;
            border: 2px solid transparent;
            background: #F0F4FA;
            color: #1F2937;
            transition: all 0.15s;
        }
        [data-testid="stSidebar"] .stButton > button:hover {
            background: #DBEAFE;
            border-color: #0056A3;
            color: #0056A3;
        }
        /* Botón activo (tipo primary = página seleccionada) */
        [data-testid="stSidebar"] .stButton > button[kind="primary"] {
            background: #0056A3 !important;
            color: #FFFFFF !important;
            border-color: #003D7A !important;
        }
        [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
            background: #003D7A !important;
            color: #FFFFFF !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ── Botón REGISTRAR ──────────────────────────────────────────────────────
    _reg_active = st.session_state.pagina_activa == "REGISTRAR"
    _arrow = "▾" if st.session_state.registrar_submenu_open else "▸"
    _reg_lbl = f"📝 **Registrar** {_arrow}" if _reg_active else f"📝 Registrar {_arrow}"
    if st.button(
        _reg_lbl,
        key="_nav_REGISTRAR",
        use_container_width=True,
        type="primary" if _reg_active else "secondary",
    ):
        if st.session_state.pagina_activa == "REGISTRAR":
            st.session_state.registrar_submenu_open = not st.session_state.registrar_submenu_open
        else:
            st.session_state.pagina_activa = "REGISTRAR"
            st.session_state.registrar_submenu_open = True
        st.rerun()

    # ── Sub-menú inline (justo debajo de Registrar) ──────────────────────────
    if st.session_state.registrar_submenu_open:
        _tipo_opts = ["RUTAS", "TRANSUIZA", "SEGUIMIENTOS"]
        _t_icons = {"RUTAS": "🚛", "TRANSUIZA": "🏭", "SEGUIMIENTOS": "🔬"}
        st.markdown(
            "<div style='margin:4px 0 4px 10px;border-left:3px solid #0056A3;"
            "padding-left:8px;'>"
            "<span style='font-size:0.68rem;font-weight:700;color:#6B7280;"
            "letter-spacing:.06em;'>TIPO DE ANÁLISIS</span></div>",
            unsafe_allow_html=True,
        )
        for _t in _tipo_opts:
            _t_active = (
                st.session_state.pagina_activa == "REGISTRAR"
                and st.session_state.tipo_registrar == _t
            )
            _t_lbl = f"  {_t_icons[_t]} **{_t}**" if _t_active else f"  {_t_icons[_t]} {_t}"
            if st.button(
                _t_lbl,
                key=f"_subnav_{_t}",
                use_container_width=True,
                type="primary" if _t_active else "secondary",
            ):
                st.session_state.pagina_activa = "REGISTRAR"
                st.session_state.tipo_registrar = _t
                st.session_state.registrar_submenu_open = False
                st.rerun()

    # ── Botón HISTORIAL ──────────────────────────────────────────────────────
    _hist_active = st.session_state.pagina_activa == "HISTORIAL"
    _hist_lbl = "🗂️ **Historial**" if _hist_active else "🗂️ Historial"
    if st.button(
        _hist_lbl,
        key="_nav_HISTORIAL",
        use_container_width=True,
        type="primary" if _hist_active else "secondary",
    ):
        st.session_state.pagina_activa = "HISTORIAL"
        st.session_state.registrar_submenu_open = False
        st.rerun()

    # ── Botón DASHBOARD ──────────────────────────────────────────────────────
    _dash_active = st.session_state.pagina_activa == "DASHBOARD"
    _dash_lbl = "📊 **Dashboard**" if _dash_active else "📊 Dashboard"
    if st.button(
        _dash_lbl,
        key="_nav_DASHBOARD",
        use_container_width=True,
        type="primary" if _dash_active else "secondary",
    ):
        st.session_state.pagina_activa = "DASHBOARD"
        st.session_state.registrar_submenu_open = False
        st.rerun()

    st.markdown("<hr style='border-color:#E5E7EB;margin:18px 0;'>", unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:0.72rem;color:#9CA3AF;text-align:center;'>"
        f"Sección activa:<br><strong style='color:#0056A3;'>"
        f"{st.session_state.pagina_activa}</strong></div>",
        unsafe_allow_html=True,
    )


if st.session_state.pagina_activa == "REGISTRAR":
    # ── Variables de tipo desde el sidebar ───────────────────────────────────
    tipo_servicio = st.session_state.tipo_registrar
    sub_tipo_seg = st.session_state.sub_tipo_registrar
    st.session_state["_tipo_servicio_guardado"] = tipo_servicio
    st.session_state["_sub_tipo_seg_guardado"] = sub_tipo_seg

    if tipo_servicio == "RUTAS":
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;
                            margin-bottom:6px;">
                  <span style="font-size:1.35rem;">📋</span>
                  <span style="font-size:1.35rem;font-weight:700;
                               color:#0056A3;letter-spacing:.5px;
                               font-family:'Segoe UI',sans-serif;">
                    SEGUIMIENTO DE RUTAS
                  </span>
                </div>""",
            unsafe_allow_html=True,
        )

        # ── Generación de claves para reset limpio al guardar ─────────
        if "_ruta_fg" not in st.session_state:
            st.session_state._ruta_fg = 0
        _fg = st.session_state._ruta_fg

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📋 Datos de Identificación
               </div>""",
            unsafe_allow_html=True,
        )
        r1c1, r1c2 = st.columns(2)
        fecha_ruta = r1c1.date_input(
            "📅  FECHA DE LA RUTA", now_col(),
            key=f"fecha_ruta_{_fg}", format="DD/MM/YYYY",
        )
        nombre_ruta = r1c2.text_input(
            "📍  NOMBRE DE LA RUTA", placeholder="ESCRIBA AQUÍ...",
            key=f"nombre_ruta_{_fg}", on_change=sanitizar_nombre_ruta,
            args=(f"nombre_ruta_{_fg}",),
        )
        r2c1, r2c2, r2c3 = st.columns(3)
        placa = r2c1.text_input(
            "🚚  PLACA DE VEHÍCULO", placeholder="AAA000",
            key=f"placa_vehiculo_{_fg}", on_change=validar_placa,
        )
        conductor = r2c2.text_input(
            "👤  CONDUCTOR", placeholder="NOMBRE COMPLETO",
            key=f"conductor_{_fg}", on_change=convertir_a_mayusculas,
            args=(f"conductor_{_fg}",),
        )
        volumen = r2c3.number_input(
            "📦  VOLUMEN (L)", min_value=0, value=None, step=1,
            format="%d", placeholder="DIGITE VOLUMEN", key=f"volumen_ruta_{_fg}",
        )
        activar_siguiente_con_enter()

        st.markdown("---")

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 🧪 Análisis de Calidad de Ruta
               </div>""",
            unsafe_allow_html=True,
        )
        cq1, cq2 = st.columns(2)

        with cq1:
            solidos_raw = st.text_input(
                "SÓLIDOS TOTALES (%)",
                key=f"solidos_totales_{_fg}",
                placeholder="Ej: 12.80",
            )
            try:
                solidos_totales = float(solidos_raw.replace(",", ".")) if solidos_raw else None
            except ValueError:
                solidos_totales = None
                st.warning("⚠️ Ingrese un número válido")

            if solidos_totales is not None and 0 < solidos_totales < 12.60:
                st.error("🚨 ALERTA: SÓLIDOS POR DEBAJO DE 12.60%")
                st.markdown(
                    """
                    <style>
                    div[data-testid="stTextInput"]:has(input[aria-label="SÓLIDOS TOTALES (%)"]) input {
                        border: 2px solid red !important;
                        background-color: #fff0f0 !important;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
            elif solidos_totales is not None and solidos_totales >= 12.60:
                st.success("✅ Sólidos dentro del parámetro")

        with cq2:
            crioscopia_raw = st.text_input(
                "CRIOSCOPIA (°C)",
                key=f"crioscopia_{_fg}",
                value="-0.",
                placeholder="-0.530",
            )
            try:
                crioscopia = float(crioscopia_raw.replace(",", ".")) if crioscopia_raw not in ("", "-", "-0", "-0.") else None
            except ValueError:
                crioscopia = None
                st.warning("⚠️ Ingrese un número válido")

            if crioscopia is not None and crioscopia > -0.535:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MAYOR A -0.535)")
            elif crioscopia is not None and crioscopia < -0.550:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")
            elif crioscopia is not None:
                st.success("✅ Crioscopia dentro del parámetro")

        st.markdown("---")

        # ── Imágenes de Muestras ───────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📷 Imágenes de Muestras
               </div>""",
            unsafe_allow_html=True,
        )
        if "imagenes_confirmadas" not in st.session_state:
            st.session_state.imagenes_confirmadas = False
        if "imagenes_nombres_guardados" not in st.session_state:
            st.session_state.imagenes_nombres_guardados = []

        imagenes_subidas = st.file_uploader(
            "ADJUNTAR IMÁGENES DE MUESTRAS DE LA RUTA",
            type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key=f"imagenes_muestras_{_fg}",
            label_visibility="visible",
        )

        if imagenes_subidas:
            nombres_actuales = [f.name for f in imagenes_subidas]
            if nombres_actuales != st.session_state.imagenes_nombres_guardados:
                st.session_state.imagenes_confirmadas = False

            # ── Miniaturas en cuadrícula HTML (bordes redondeados) ────
            confirmed = st.session_state.imagenes_confirmadas
            thumb_html = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
            for img in imagenes_subidas:
                raw_bytes = img.read()
                b64 = base64.b64encode(raw_bytes).decode()
                ext = img.name.rsplit(".", 1)[-1].lower()
                mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
                nombre_corto = img.name if len(img.name) <= 16 else img.name[:14] + "…"
                check_html = (
                    "<div style='color:#16a34a;font-size:12px;"
                    "text-align:center;font-weight:600;'>✅ Guardada</div>"
                    if confirmed else
                    f"<div style='font-size:10px;color:#888;text-align:center;'>{nombre_corto}</div>"
                )
                border_color = "#16a34a" if confirmed else "#D1D5DB"
                thumb_html += (
                    f"<div style='display:flex;flex-direction:column;"
                    f"align-items:center;gap:4px;'>"
                    f"<img src='data:{mime};base64,{b64}' "
                    f"style='width:150px;height:150px;object-fit:cover;"
                    f"border-radius:10px;border:2px solid {border_color};"
                    f"box-shadow:0 2px 6px rgba(0,0,0,0.08);background:#F4F4F4;'/>"
                    f"{check_html}</div>"
                )
                img.seek(0)  # reset cursor para uso posterior
            thumb_html += "</div>"
            st.markdown(thumb_html, unsafe_allow_html=True)

            # ── Botón guardar / confirmación ───────────────────────
            if not st.session_state.imagenes_confirmadas:
                st.markdown("<div style='margin-top:8px;'></div>",
                            unsafe_allow_html=True)
                if st.button("💾 GUARDAR IMÁGENES",
                             use_container_width=False):
                    st.session_state.imagenes_confirmadas = True
                    st.session_state.imagenes_nombres_guardados = nombres_actuales
                    st.rerun()
            else:
                st.success("✅ Imágenes guardadas correctamente.")
        else:
            st.session_state.imagenes_confirmadas = False
            st.caption("No se han adjuntado imágenes.")

        st.markdown("---")

        if "estaciones_guardadas" not in st.session_state:
            st.session_state.estaciones_guardadas = []
        if "form_ver" not in st.session_state:
            st.session_state.form_ver = 0

        def parse_num(val, default=None):
            if val is None:
                return default
            try:
                return float(str(val).replace(",", "."))
            except ValueError:
                return default

        # ── Formulario nueva estación ──────────────────────────────────
        with st.container(border=True):
            v = st.session_state.form_ver
            num_nueva = len(st.session_state.estaciones_guardadas) + 1

            # Lookup nombre desde catálogo con el código actual en session_state
            _cat_r = load_catalogo()
            _cat_r_cod = dict(zip(_cat_r["codigo"], _cat_r["nombre"]))
            _cod_actual = st.session_state.get(f"nue_codigo_{v}", "").strip()
            _nom_actual = _cat_r_cod.get(_cod_actual) or _cat_r_cod.get(_cod_actual.upper(), "")
            if _nom_actual:
                st.markdown(
                    f"**Agregar Estación — #{num_nueva}"
                    f"&nbsp;&nbsp;<span style='color:#0056A3;font-size:1em;'>"
                    f"· {_nom_actual}</span>**",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(f"**Agregar Estación — #{num_nueva}**")

            if f"nue_crio_{v}" not in st.session_state:
                st.session_state[f"nue_crio_{v}"] = "-0."

            f1, f2, f3, f4, f5, f6 = st.columns([1.5, 1, 1, 1, 1.5, 1])
            form_codigo   = f1.text_input("CÓDIGO", key=f"nue_codigo_{v}",
                                          placeholder="CÓDIGO",
                                          on_change=convertir_a_mayusculas,
                                          args=(f"nue_codigo_{v}",))
            form_grasa    = f2.number_input("GRASA (%)", key=f"nue_grasa_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_solidos  = f3.number_input("SÓL. TOT. (%)", key=f"nue_solidos_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_proteina = f4.number_input("PROTEÍNA (%)", key=f"nue_proteina_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_crio_raw = f5.text_input("CRIOSCOPIA (°C)", key=f"nue_crio_{v}",
                                          placeholder="-0.530")
            form_vol      = f6.number_input("VOLUMEN (L)", key=f"nue_vol_{v}",
                                            min_value=0, step=1,
                                            value=None, placeholder="0")

            form_crio_val = (parse_num(form_crio_raw)
                             if form_crio_raw not in ("", "-", "-0", "-0.")
                             else None)

            if form_solidos is not None and 0 < form_solidos < 12.60:
                st.error("🚨 SÓLIDOS POR DEBAJO DE 12.60%")

            form_agua_pct = None
            if form_crio_val is not None and form_crio_val > -0.530:
                aw1, aw2 = st.columns([2, 1])
                aw1.warning("💧 ALERTA: PRESENCIA DE AGUA — CRIOSCOPIA MAYOR A -0.530")
                form_agua_pct = aw2.number_input(
                    "% AGUA AÑADIDA", key=f"nue_agua_{v}",
                    min_value=0.0, max_value=100.0, step=0.1,
                    format="%.1f", value=None, placeholder="0.0")
            elif form_crio_val is not None and form_crio_val < -0.550:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")

            q1, q2, q3, q4 = st.columns([0.6, 0.6, 0.8, 2])
            form_alcohol  = q1.selectbox("ALCOHOL", options=["N/A", "+", "-"],
                                         key=f"nue_alcohol_{v}")
            form_cloruros = q2.selectbox("CLORUROS", options=["N/A", "+", "-"],
                                         key=f"nue_cloruros_{v}")
            form_neutral  = q3.selectbox("NEUTRALIZANTES", options=["N/A", "+", "-"],
                                         key=f"nue_neutral_{v}")
            form_obs = q4.text_input("OBSERVACIONES", key=f"nue_obs_{v}",
                                     placeholder="Ingrese observaciones...")

            if st.button("💾 GUARDAR", type="primary",
                         use_container_width=True):
                st.session_state.estaciones_guardadas.append({
                    "codigo":         form_codigo,
                    "grasa":          form_grasa,
                    "solidos":        form_solidos,
                    "proteina":       form_proteina,
                    "crioscopia":     form_crio_raw if form_crio_val is not None else None,
                    "volumen":        form_vol,
                    "alcohol":        form_alcohol,
                    "cloruros":       form_cloruros,
                    "neutralizantes": form_neutral,
                    "agua_pct":       form_agua_pct,
                    "obs":            form_obs,
                })
                st.session_state.form_ver += 1
                st.rerun()


        st.markdown("---")

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📦 Calidad por Estación
               </div>""",
            unsafe_allow_html=True,
        )

        # ── Data editor ────────────────────────────────────────────────
        EDITOR_COLS = ["codigo", "grasa", "solidos", "proteina", "crioscopia",
                       "agua_pct", "volumen", "alcohol", "cloruros",
                       "neutralizantes", "obs"]

        if st.session_state.estaciones_guardadas:
            df_est = pd.DataFrame(st.session_state.estaciones_guardadas,
                                  columns=EDITOR_COLS)
        else:
            df_est = pd.DataFrame(columns=EDITOR_COLS)

        for c in ["grasa", "solidos", "proteina", "agua_pct"]:
            df_est[c] = pd.to_numeric(df_est[c], errors="coerce")
        df_est["volumen"] = pd.to_numeric(df_est["volumen"],
                                          errors="coerce").astype("Int64")

        # Columna NOMBRE ESTACIÓN derivada del catálogo (solo lectura)
        _cat_tab = load_catalogo()
        _cat_tab_map = dict(zip(_cat_tab["codigo"], _cat_tab["nombre"]))
        df_est["nombre_estacion"] = df_est["codigo"].apply(
            lambda c: _cat_tab_map.get(str(c).strip(), "")
                      if pd.notna(c) else ""
        )

        _nv = st.session_state.get("_est_nombre_ver", 0)
        edited = st.data_editor(
            df_est,
            num_rows="dynamic",
            use_container_width=True,
            key=f"de_est_{st.session_state.form_ver}_{_nv}",
            column_config={
                "codigo":        st.column_config.TextColumn("CÓDIGO"),
                "grasa":         st.column_config.NumberColumn(
                                     "GRASA (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "solidos":       st.column_config.NumberColumn(
                                     "SÓL.TOT. (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "proteina":      st.column_config.NumberColumn(
                                     "PROTEÍNA (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "crioscopia":    st.column_config.TextColumn("CRIOSCOPIA (°C)"),
                "volumen":       st.column_config.NumberColumn(
                                     "VOLUMEN (L)", format="%d",
                                     min_value=0, step=1),
                "alcohol":       st.column_config.SelectboxColumn(
                                     "ALCOHOL", options=["N/A", "+", "-"],
                                     required=True),
                "cloruros":      st.column_config.SelectboxColumn(
                                     "CLORUROS", options=["N/A", "+", "-"],
                                     required=True),
                "neutralizantes":st.column_config.SelectboxColumn(
                                     "NEUTRALIZANTES", options=["N/A", "+", "-"],
                                     required=True),
                "agua_pct":      st.column_config.NumberColumn(
                                     "% AGUA", format="%.1f",
                                     min_value=0.0, max_value=100.0),
                "obs":           st.column_config.TextColumn("OBSERVACIONES"),
                "nombre_estacion": st.column_config.TextColumn(
                                     "NOMBRE ESTACIÓN", disabled=True),
            },
            hide_index=True,
        )

        # Sincronizar ediciones/eliminaciones de vuelta al estado (sin nombre_estacion)
        _prev_codes = list(df_est["codigo"].fillna("").astype(str).str.strip())
        raw = json.loads(edited.to_json(orient="records"))
        st.session_state.estaciones_guardadas = [
            {k: v for k, v in r.items() if k != "nombre_estacion"}
            for r in raw
            if any(v is not None and str(v).strip() != ""
                   for k, v in r.items() if k != "nombre_estacion")
        ]
        _new_codes = [str(r.get("codigo") or "").strip()
                      for r in st.session_state.estaciones_guardadas]
        if _prev_codes != _new_codes:
            st.session_state["_est_nombre_ver"] = st.session_state.get("_est_nombre_ver", 0) + 1
            st.rerun()

        st.markdown("---")
        # ── Reconciliación de volúmenes ────────────────────────────────
        st.markdown("---")
        vol_ruta = volumen if volumen is not None else 0
        vol_est_total = 0
        for e in st.session_state.estaciones_guardadas:
            v_e = e.get("volumen")
            if v_e is not None:
                try:
                    vol_est_total += int(v_e)
                except (ValueError, TypeError):
                    pass

        col_res1, col_res2, col_res3 = st.columns(3)
        col_res1.metric("VOLUMEN DECLARADO DE RUTA (L)",
                        f"{int(vol_ruta):,}" if vol_ruta else "—")
        col_res2.metric("VOLUMEN SUMA ESTACIONES (L)",
                        f"{int(vol_est_total):,}" if vol_est_total else "—")
        diferencia = vol_est_total - vol_ruta if vol_ruta else 0
        col_res3.metric("DIFERENCIA (L)",
                        f"{int(diferencia):+,}" if vol_ruta else "—")

        if vol_ruta and vol_est_total and vol_ruta != vol_est_total:
            st.warning(
                f"⚠️ El volumen de estaciones ({int(vol_est_total):,} L) no coincide "
                f"con el volumen declarado de la ruta ({int(vol_ruta):,} L). "
                f"Diferencia: {int(diferencia):+,} L"
            )
        elif vol_ruta and vol_est_total and vol_ruta == vol_est_total:
            st.success("✅ El volumen de estaciones coincide con el volumen de la ruta.")

        # ── Ponderados a nivel de ruta (usados en exportación y guardado) ────
        _ests = st.session_state.estaciones_guardadas
        _pond_st, _pond_ic = [], []
        for _e in _ests:
            _v  = parse_num(_e.get("volumen"))
            _s  = parse_num(_e.get("solidos"))
            _cr = parse_num(_e.get("crioscopia"))
            _pond_st.append(round(_v * _s,  2) if _v is not None and _s  is not None else None)
            _pond_ic.append(round(_v * _cr, 3) if _v is not None and _cr is not None else None)
        _vol_total = sum(parse_num(_e.get("volumen")) or 0 for _e in _ests)
        _st_pond = round(sum(x for x in _pond_st if x is not None) / _vol_total, 2) if _vol_total else None
        _ic_pond = round(sum(x for x in _pond_ic if x is not None) / _vol_total, 3) if _vol_total else None

        # ── GUARDAR RUTA EN HISTORIAL ──────────────────────────────────
        st.markdown("---")
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 💾 Guardar en Historial
               </div>""",
            unsafe_allow_html=True,
        )

        if "ruta_guardada_ok" not in st.session_state:
            st.session_state.ruta_guardada_ok = False

        if st.button(
            "💾  GUARDAR RUTA",
            type="primary",
            use_container_width=True,
            key="btn_guardar_ruta",
        ):
            _fotos_prefix = f"RUTAS_{(placa or 'SIN_PLACA').upper()}"
            _fotos_paths  = save_fotos_to_disk(imagenes_subidas or [], _fotos_prefix)
            save_ruta_to_csv({
                "tipo_seguimiento": "RUTAS",
                "fecha":            fecha_ruta.strftime("%d/%m/%Y") if fecha_ruta else "",
                "ruta":             nombre_ruta or "",
                "placa":            placa or "",
                "conductor":        conductor or "",
                "volumen_declarado": int(volumen) if volumen else "",
                "vol_estaciones":   int(vol_est_total) if vol_est_total else "",
                "diferencia":       int(diferencia) if vol_ruta else "",
                "solidos_ruta":     solidos_totales if solidos_totales is not None else "",
                "crioscopia_ruta":  crioscopia if crioscopia is not None else "",
                "st_pond":          _st_pond if _st_pond is not None else "",
                "ic_pond":          _ic_pond if _ic_pond is not None else "",
                "num_estaciones":   len(st.session_state.estaciones_guardadas),
                "guardado_en":      now_col().strftime("%d/%m/%Y %H:%M"),
                "estaciones_json":  json.dumps(
                    st.session_state.estaciones_guardadas, ensure_ascii=False
                ),
                "fotos_json":       json.dumps(_fotos_paths, ensure_ascii=False),
            })
            # ── Limpiar todos los campos para la siguiente ruta ────────
            # Incrementar generación → todos los widgets de identificación
            # y calidad obtienen una clave nueva y se renderizan vacíos
            st.session_state._ruta_fg = (
                st.session_state.get("_ruta_fg", 0) + 1
            )
            # Limpiar campos de estación (nue_*)
            for _k in list(st.session_state.keys()):
                if _k.startswith("nue_"):
                    st.session_state.pop(_k, None)
            st.session_state.pop("imagenes_muestras", None)
            st.session_state.estaciones_guardadas          = []
            st.session_state.form_ver                      = (
                st.session_state.get("form_ver", 0) + 1
            )
            st.session_state.imagenes_confirmadas          = False
            st.session_state.imagenes_nombres_guardados    = []
            st.session_state.ruta_guardada_ok              = True
            clear_draft_state()
            st.rerun()

        if st.session_state.ruta_guardada_ok:
            st.success(
                "✅ Ruta guardada exitosamente en el historial. "
                "Puedes consultarla en **📊 Historial de Rutas** al final de la página."
            )
            st.session_state.ruta_guardada_ok = False

    elif tipo_servicio == "TRANSUIZA":
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                 <span style="font-size:1.35rem;">🚛</span>
                 <span style="font-size:1.35rem;font-weight:700;color:#0056A3;
                              letter-spacing:.5px;font-family:'Segoe UI',sans-serif;">
                   SEGUIMIENTO TRANSUIZA
                 </span>
               </div>""",
            unsafe_allow_html=True,
        )

        # ── Datos de identificación ────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📋 Datos de Identificación
               </div>""",
            unsafe_allow_html=True,
        )
        with st.container(border=True):
            tc1, tc2, tc3 = st.columns(3)
            trans_fecha = tc1.date_input(
                "📅 FECHA", now_col(), key="trans_fecha", format="DD/MM/YYYY"
            )
            trans_placa = tc2.text_input(
                "🚚 PLACA DEL VEHÍCULO", placeholder="AAA000", key="trans_placa",
                on_change=lambda: st.session_state.__setitem__(
                    "trans_placa",
                    re.sub(r"[^A-Z0-9]", "", st.session_state.get("trans_placa", "").upper())
                ),
            )
            trans_st_carrotanque = tc3.number_input(
                "🏷️ ST DEL CARROTANQUE (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key="trans_st_carrotanque",
            )
        activar_siguiente_con_enter()

        # ── Calidad de la muestra ──────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 🧪 Calidad de la Muestra
               </div>""",
            unsafe_allow_html=True,
        )
        with st.container(border=True):
            qc1, qc2, qc3 = st.columns(3)
            trans_grasa = qc1.number_input(
                "GRASA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key="trans_grasa",
            )
            trans_st_muestra = qc2.number_input(
                "ST MUESTRA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key="trans_st_muestra",
            )
            trans_proteina = qc3.number_input(
                "PROTEÍNA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key="trans_proteina",
            )

            # ── Diferencia de Sólidos (automática) ────────────────────
            if trans_st_carrotanque is not None and trans_st_muestra is not None:
                dif_solidos = round(trans_st_carrotanque - trans_st_muestra, 2)
                color_dif = "#9C0006" if abs(dif_solidos) > 0.5 else "#006100"
                st.markdown(
                    f"""<div style="margin-top:12px;padding:12px 16px;
                        background:#F8FAFC;border-radius:10px;
                        border:1.5px solid #D1D5DB;text-align:center;">
                        <div style="font-size:11px;font-weight:600;color:#6B7280;
                                    letter-spacing:.4px;margin-bottom:4px;">
                            DIFERENCIA DE SÓLIDOS (ST Carrotanque − ST Muestra)
                        </div>
                        <div style="font-size:2rem;font-weight:800;color:{color_dif};">
                            {dif_solidos:+.2f} %
                        </div>
                    </div>""",
                    unsafe_allow_html=True,
                )
            else:
                dif_solidos = None
                st.info("💡 Ingrese ST del Carrotanque y ST Muestra para calcular la diferencia.")

        activar_siguiente_con_enter()

        # ── Imágenes de Muestras (TRANSUIZA) ──────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📷 Imágenes de Muestras
               </div>""",
            unsafe_allow_html=True,
        )
        if "trans_imagenes_confirmadas" not in st.session_state:
            st.session_state.trans_imagenes_confirmadas = False
        if "trans_imagenes_nombres_guardados" not in st.session_state:
            st.session_state.trans_imagenes_nombres_guardados = []
        _trans_fg = st.session_state.get("_trans_fg", 0)

        trans_imagenes_subidas = st.file_uploader(
            "ADJUNTAR IMÁGENES DE MUESTRAS TRANSUIZA",
            type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key=f"trans_imagenes_muestras_{_trans_fg}",
            label_visibility="visible",
        )

        if trans_imagenes_subidas:
            _t_nombres = [f.name for f in trans_imagenes_subidas]
            if _t_nombres != st.session_state.trans_imagenes_nombres_guardados:
                st.session_state.trans_imagenes_confirmadas = False

            _t_confirmed = st.session_state.trans_imagenes_confirmadas
            _t_thumb_html = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
            for _timg in trans_imagenes_subidas:
                _t_raw = _timg.read()
                _t_b64 = base64.b64encode(_t_raw).decode()
                _t_ext = _timg.name.rsplit(".", 1)[-1].lower()
                _t_mime = "image/jpeg" if _t_ext in ("jpg", "jpeg") else "image/png"
                _t_nombre_corto = _timg.name if len(_timg.name) <= 16 else _timg.name[:14] + "…"
                _t_check_html = (
                    "<div style='color:#16a34a;font-size:12px;"
                    "text-align:center;font-weight:600;'>✅ Guardada</div>"
                    if _t_confirmed else
                    f"<div style='font-size:10px;color:#888;text-align:center;'>{_t_nombre_corto}</div>"
                )
                _t_border = "#16a34a" if _t_confirmed else "#D1D5DB"
                _t_thumb_html += (
                    f"<div style='display:flex;flex-direction:column;"
                    f"align-items:center;gap:4px;'>"
                    f"<img src='data:{_t_mime};base64,{_t_b64}' "
                    f"style='width:150px;height:150px;object-fit:cover;"
                    f"border-radius:10px;border:2px solid {_t_border};"
                    f"box-shadow:0 2px 6px rgba(0,0,0,0.08);background:#F4F4F4;'/>"
                    f"{_t_check_html}</div>"
                )
                _timg.seek(0)
            _t_thumb_html += "</div>"
            st.markdown(_t_thumb_html, unsafe_allow_html=True)

            if not st.session_state.trans_imagenes_confirmadas:
                st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
                if st.button("💾 GUARDAR IMÁGENES", key="btn_trans_guardar_imgs",
                             use_container_width=False):
                    st.session_state.trans_imagenes_confirmadas = True
                    st.session_state.trans_imagenes_nombres_guardados = _t_nombres
                    st.rerun()
            else:
                st.success("✅ Imágenes guardadas correctamente.")
        else:
            st.session_state.trans_imagenes_confirmadas = False
            st.caption("No se han adjuntado imágenes.")

        # ── Guardar Transuiza ──────────────────────────────────────────
        st.markdown("---")
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 💾 Guardar en Historial
               </div>""",
            unsafe_allow_html=True,
        )
        if "trans_guardado_ok" not in st.session_state:
            st.session_state.trans_guardado_ok = False

        if st.button("💾  GUARDAR TRANSUIZA", type="primary",
                     use_container_width=True, key="btn_guardar_trans"):
            if not trans_placa:
                st.warning("⚠️ Ingrese la placa del vehículo.")
            else:
                _t_fotos_prefix = f"TRANS_{(trans_placa or 'SIN_PLACA').upper()}"
                _t_fotos_paths  = save_fotos_to_disk(
                    trans_imagenes_subidas or [], _t_fotos_prefix
                )
                save_ruta_to_csv({
                    "tipo_seguimiento": "TRANSUIZA",
                    "fecha":            trans_fecha.strftime("%d/%m/%Y") if trans_fecha else "",
                    "ruta":             "ENTRERIOS",
                    "placa":            (trans_placa or "").upper(),
                    "st_carrotanque":   trans_st_carrotanque if trans_st_carrotanque is not None else "",
                    "solidos_ruta":     trans_st_muestra if trans_st_muestra is not None else "",
                    "grasa_muestra":    trans_grasa if trans_grasa is not None else "",
                    "proteina_muestra": trans_proteina if trans_proteina is not None else "",
                    "diferencia_solidos": dif_solidos if dif_solidos is not None else "",
                    "guardado_en":      now_col().strftime("%d/%m/%Y %H:%M"),
                    "fotos_json":       json.dumps(_t_fotos_paths, ensure_ascii=False),
                })
                for _k in ["trans_placa", "trans_st_carrotanque", "trans_grasa",
                            "trans_st_muestra", "trans_proteina", "trans_fecha"]:
                    st.session_state.pop(_k, None)
                st.session_state["_trans_fg"]                     = _trans_fg + 1
                st.session_state.trans_imagenes_confirmadas       = False
                st.session_state.trans_imagenes_nombres_guardados = []
                st.session_state.trans_guardado_ok = True
                clear_draft_state()
                st.rerun()

        if st.session_state.trans_guardado_ok:
            st.success("✅ Registro TRANSUIZA guardado en el historial.")
            st.session_state.trans_guardado_ok = False

    elif tipo_servicio == "SEGUIMIENTOS":

        # ── Encabezado ────────────────────────────────────────────────────────
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                  <span style="font-size:1.35rem;">🔬</span>
                  <span style="font-size:1.35rem;font-weight:700;color:#0056A3;
                               letter-spacing:.5px;font-family:'Segoe UI',sans-serif;">
                    SEGUIMIENTOS
                  </span>
                </div>""",
            unsafe_allow_html=True,
        )

        # ── Pestañas de sub-tipo ─────────────────────────────────────────────
        _seg_sub_vals  = ["ESTACIONES", "ACOMPAÑAMIENTOS", "CONTRAMUESTRAS SOLICITADAS"]
        _seg_tab_icons = {"ESTACIONES": "🏭",
                          "ACOMPAÑAMIENTOS": "👥", "CONTRAMUESTRAS SOLICITADAS": "🧪"}
        _seg_tab_labels = [
            "🏭 ESTACIONES",
            "👥 ACOMPAÑAMIENTOS", "🧪 CONTRAMUESTRAS",
        ]
        _seg_tabs = st.tabs(_seg_tab_labels)

        for _ti, (_tab_ctx, _sub) in enumerate(zip(_seg_tabs, _seg_sub_vals)):
            with _tab_ctx:

                # ── Datos de Identificación ───────────────────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         📋 Datos de Identificación
                       </div>""",
                    unsafe_allow_html=True,
                )
                with st.container(border=True):
                    if _sub == "ESTACIONES":
                        _cat_df = load_catalogo()
                        _cat_map_cod = dict(zip(_cat_df["codigo"], _cat_df["nombre"]))
                        _cat_map_nom = dict(zip(_cat_df["nombre"], _cat_df["codigo"]))

                        # callbacks de llenado cruzado
                        def _fill_nombre_from_cod(_ti=_ti, _m=_cat_map_cod):
                            cod = st.session_state.get(f"seg_codigo_{_ti}", "").strip()
                            nom = _m.get(cod) or _m.get(cod.upper())
                            if nom:
                                st.session_state[f"seg_nombre_{_ti}"] = nom

                        def _fill_cod_from_nombre(_ti=_ti, _m=_cat_map_nom):
                            raw = st.session_state.get(f"seg_nombre_{_ti}", "")
                            nom = raw.upper()
                            st.session_state[f"seg_nombre_{_ti}"] = nom
                            cod = _m.get(nom.strip())
                            if cod:
                                st.session_state[f"seg_codigo_{_ti}"] = cod

                        sid1, sid2, sid3 = st.columns([1, 1, 2])
                        seg_fecha = sid1.date_input(
                            "📅 FECHA", now_col(),
                            key=f"seg_fecha_{_ti}", format="DD/MM/YYYY",
                        )
                        seg_codigo = sid2.text_input(
                            "🔖 CÓDIGO", placeholder="Ej: 6085",
                            key=f"seg_codigo_{_ti}",
                            on_change=_fill_nombre_from_cod,
                        )
                        st.session_state.setdefault(f"seg_nombre_{_ti}", "")
                        sid3.text_input(
                            "📍 NOMBRE ESTACIÓN", placeholder="Ej: LUCERITO",
                            key=f"seg_nombre_{_ti}",
                            on_change=_fill_cod_from_nombre,
                        )
                        activar_siguiente_con_enter()

                        seg_quien_trajo = ""
                        seg_ruta_acomp  = ""
                        seg_responsable = ""

                    elif _sub == "ACOMPAÑAMIENTOS":
                        sa1, sa2 = st.columns(2)
                        seg_fecha = sa1.date_input(
                            "📅 FECHA", now_col(),
                            key=f"seg_fecha_{_ti}", format="DD/MM/YYYY",
                        )
                        seg_ruta_acomp = sa2.text_input(
                            "📍 NOMBRE DE LA RUTA", placeholder="ESCRIBA AQUÍ...",
                            key=f"seg_ruta_acomp_{_ti}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_ruta_acomp_{_ti}",),
                        )
                        sb1, sb2 = st.columns(2)
                        seg_quien_trajo = sb1.text_input(
                            "👤 ENTREGADO POR", placeholder="NOMBRE COMPLETO",
                            key=f"seg_quien_trajo_{_ti}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_quien_trajo_{_ti}",),
                        )
                        seg_vol_declarado_id = sb2.number_input(
                            "🪣 VOLUMEN (L)", min_value=0, step=1,
                            value=None, placeholder="DIGITE VOLUMEN",
                            key=f"seg_vol_declarado_{_ti}",
                        )
                        seg_codigo      = ""
                        seg_responsable = ""

                    else:  # CONTRAMUESTRAS SOLICITADAS
                        sc1, sc2 = st.columns(2)
                        seg_fecha = sc1.date_input(
                            "📅 FECHA DE LAS MUESTRAS", now_col(),
                            key=f"seg_fecha_{_ti}", format="DD/MM/YYYY",
                        )
                        seg_responsable = sc2.text_input(
                            "👤 ENTREGADO POR", placeholder="NOMBRE...",
                            key=f"seg_responsable_{_ti}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_responsable_{_ti}",),
                        )
                        seg_codigo      = ""
                        seg_quien_trajo = ""
                        seg_ruta_acomp  = ""

                activar_siguiente_con_enter()

                # ── ACOMPAÑAMIENTOS: sección análisis justo tras identificación ─
                if _sub == "ACOMPAÑAMIENTOS":
                    st.markdown("---")
                    st.markdown(
                        """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                       margin:10px 0 6px 0;letter-spacing:.4px;
                                       border-left:4px solid #0056A3;padding-left:10px;">
                             🔬 Análisis de Calidad de Ruta
                           </div>""",
                        unsafe_allow_html=True,
                    )
                    _aq1, _aq2 = st.columns(2)
                    seg_solidos_ruta = _aq1.number_input(
                        "SÓLIDOS TOTALES (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None, placeholder="Ej: 12.80",
                        key=f"seg_solidos_ruta_{_ti}",
                    )
                    seg_crios_raw = _aq2.text_input(
                        "CRIOSCOPIA (°C)", value="-0.", placeholder="-0.530",
                        key=f"seg_crios_raw_{_ti}",
                    )
                    try:
                        seg_crioscopia_ruta = float(seg_crios_raw.replace(",", ".")) if seg_crios_raw.strip() else None
                    except Exception:
                        seg_crioscopia_ruta = None
                    st.markdown("---")

                # ── Parámetros de Calidad ─────────────────────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         🧪 Parámetros de Calidad
                       </div>""",
                    unsafe_allow_html=True,
                )
                _qk = st.session_state.get(f"seg_quality_key_counter_{_ti}", 0)
                with st.container(border=True):
                    seg_id_muestra = ""
                    seg_volumen    = None
                    if _sub in ("ACOMPAÑAMIENTOS", "CONTRAMUESTRAS SOLICITADAS"):
                        _cat_df_m  = load_catalogo()
                        _cat_m_cod = dict(zip(_cat_df_m["codigo"], _cat_df_m["nombre"]))
                        _cat_m_nom = dict(zip(_cat_df_m["nombre"], _cat_df_m["codigo"]))

                        def _fill_nom_m(_ti=_ti, _qk=_qk, _m=_cat_m_cod):
                            cod = st.session_state.get(f"seg_id_muestra_{_ti}_{_qk}", "").strip()
                            st.session_state[f"seg_id_muestra_{_ti}_{_qk}"] = cod.upper()
                            nom = _m.get(cod) or _m.get(cod.upper())
                            if nom:
                                st.session_state[f"seg_nom_muestra_{_ti}_{_qk}"] = nom

                        def _fill_cod_m(_ti=_ti, _qk=_qk, _m=_cat_m_nom):
                            raw = st.session_state.get(f"seg_nom_muestra_{_ti}_{_qk}", "")
                            nom = raw.upper()
                            st.session_state[f"seg_nom_muestra_{_ti}_{_qk}"] = nom
                            cod = _m.get(nom.strip())
                            if cod:
                                st.session_state[f"seg_id_muestra_{_ti}_{_qk}"] = cod

                        if _sub == "ACOMPAÑAMIENTOS":
                            _idc1, _idc2, _idc3 = st.columns([1, 2, 1])
                        else:
                            _idc1, _idc2 = st.columns([1, 2])
                        seg_id_muestra = _idc1.text_input(
                            "🔖 CÓDIGO", placeholder="Ej: 6085",
                            key=f"seg_id_muestra_{_ti}_{_qk}",
                            on_change=_fill_nom_m,
                        )
                        st.session_state.setdefault(f"seg_nom_muestra_{_ti}_{_qk}", "")
                        _idc2.text_input(
                            "📍 NOMBRE ESTACIÓN", placeholder="Ej: LUCERITO",
                            key=f"seg_nom_muestra_{_ti}_{_qk}",
                            on_change=_fill_cod_m,
                        )
                        if _sub == "ACOMPAÑAMIENTOS":
                            seg_volumen = _idc3.number_input(
                                "VOLUMEN (L)", min_value=0, step=1,
                                value=None, placeholder="0",
                                key=f"seg_volumen_{_ti}_{_qk}",
                            )
                        else:
                            seg_volumen = None
                        activar_siguiente_con_enter()

                    sq1, sq2, sq3, sq4 = st.columns(4)
                    seg_grasa = sq1.number_input(
                        "GRASA (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_grasa_{_ti}_{_qk}",
                    )
                    seg_st = sq2.number_input(
                        "ST (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_st_{_ti}_{_qk}",
                    )
                    seg_proteina = sq4.number_input(
                        "PROTEÍNA (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_proteina_{_ti}_{_qk}",
                    )
                    with sq3:
                        seg_ic_raw = st.text_input(
                            "IC (°C)", key=f"seg_ic_raw_{_ti}_{_qk}",
                            value="-0.", placeholder="-0.530",
                        )
                        try:
                            seg_ic = float(seg_ic_raw.replace(",", ".")) \
                                if seg_ic_raw not in ("", "-", "-0", "-0.") else None
                        except ValueError:
                            seg_ic = None
                            st.warning("⚠️ Ingrese un número válido")
                    _ic_fuera = seg_ic is not None and seg_ic > -0.530
                    _ic_bajo  = seg_ic is not None and seg_ic < -0.550
                    seg_agua = None
                    if _ic_fuera:
                        _aw1, _aw2 = st.columns([1, 2])
                        with _aw1:
                            seg_agua = st.number_input(
                                "💧 AGUA ADICIONADA (%)",
                                min_value=0.0, max_value=100.0,
                                step=0.01, format="%.2f", value=None,
                                placeholder="0.00", key=f"seg_agua_{_ti}_{_qk}",
                            )

                    if seg_st is not None and seg_st > 0 and seg_st < 12.60:
                        st.error("🚨 ALERTA: ST FUERA DE RANGO (MENOR A 12.60%)")
                    elif seg_st is not None and seg_st > 0:
                        st.success("✅ ST dentro del parámetro")

                    if _ic_fuera:
                        st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MAYOR A -0.530)")
                    elif _ic_bajo:
                        st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")
                    elif seg_ic is not None:
                        st.success("✅ Crioscopia dentro del parámetro")

                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                    sq5, sq6, sq7 = st.columns(3)
                    opciones_tri = ["N/A", "NEGATIVO (−)", "POSITIVO (+)"]
                    seg_alcohol        = sq5.selectbox("ALCOHOL",        opciones_tri, key=f"seg_alcohol_{_ti}_{_qk}")
                    seg_cloruros       = sq6.selectbox("CLORUROS",       opciones_tri, key=f"seg_cloruros_{_ti}_{_qk}")
                    seg_neutralizantes = sq7.selectbox("NEUTRALIZANTES", opciones_tri, key=f"seg_neutralizantes_{_ti}_{_qk}")

                    _positivos = [p for p, v in [
                        ("ALCOHOL", seg_alcohol), ("CLORUROS", seg_cloruros),
                        ("NEUTRALIZANTES", seg_neutralizantes),
                    ] if v == "POSITIVO (+)"]
                    if _positivos:
                        st.error(f"🚨 ALERTA: {', '.join(_positivos)} POSITIVO(S) — ADULTERACIÓN")

                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                    seg_observaciones = st.text_area(
                        "📝 OBSERVACIONES", placeholder="ESCRIBA AQUÍ...",
                        key=f"seg_observaciones_{_ti}_{_qk}", height=90,
                    )

                activar_siguiente_con_enter()

                # ── ACOMPAÑAMIENTOS: agregar muestra ─────────────────────
                if _sub == "ACOMPAÑAMIENTOS":
                    _acomp_key = "acomp_muestras"
                    if _acomp_key not in st.session_state:
                        st.session_state[_acomp_key] = []
                    if st.button("➕  AGREGAR MUESTRA", use_container_width=True,
                                 key=f"btn_agregar_muestra_{_ti}"):
                        st.session_state[_acomp_key].append({
                            "ID": seg_id_muestra or "",
                            "VOLUMEN (L)": int(seg_volumen) if seg_volumen is not None else "",
                            "GRASA (%)": f"{seg_grasa:.2f}" if seg_grasa is not None else "",
                            "ST (%)":    f"{seg_st:.2f}"    if seg_st    is not None else "",
                            "IC (°C)":   f"{seg_ic:.3f}"    if seg_ic    is not None else "",
                            "AGUA (%)":  f"{seg_agua:.2f}"  if seg_agua  is not None else "",
                            "ALCOHOL":   seg_alcohol, "CLORUROS": seg_cloruros,
                            "NEUTRALIZANTES": seg_neutralizantes,
                            "OBS":       seg_observaciones or "",
                            "_volumen": seg_volumen,
                            "_grasa": seg_grasa, "_st": seg_st, "_ic": seg_ic,
                            "_proteina": seg_proteina,
                            "_agua": seg_agua, "_alcohol": seg_alcohol,
                            "_cloruros": seg_cloruros, "_neutralizantes": seg_neutralizantes,
                            "_obs": seg_observaciones or "",
                        })
                        st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                        st.rerun()
                    if st.session_state[_acomp_key]:
                        st.markdown(
                            f"""<div style="font-size:0.9rem;font-weight:700;color:#0056A3;
                                           margin:10px 0 4px 0;">
                                 📋 {len(st.session_state[_acomp_key])} muestra(s) registrada(s)
                               </div>""",
                            unsafe_allow_html=True,
                        )
                        _cat_pa  = load_catalogo()
                        _cat_pa_map = dict(zip(_cat_pa["codigo"], _cat_pa["nombre"]))
                        _rows_pa = []
                        for m in st.session_state[_acomp_key]:
                            _cod_pa = str(m.get("ID", "") or "").strip()
                            _nom_pa = _cat_pa_map.get(_cod_pa, "")
                            _row_pa = {"ID": _cod_pa, "NOMBRE ESTACIÓN": _nom_pa}
                            _row_pa.update({k: v for k, v in m.items()
                                            if not k.startswith("_") and k != "ID"})
                            _rows_pa.append(_row_pa)
                        df_prev_a = pd.DataFrame(_rows_pa)
                        st.dataframe(df_prev_a, use_container_width=True, hide_index=True)

                # ── ACOMPAÑAMIENTOS: calcular ponderados para guardado ────
                if _sub == "ACOMPAÑAMIENTOS":
                    seg_vol_declarado = seg_vol_declarado_id
                    _acomp_vols_f = [
                        (lambda v: v if v is not None else None)(
                            (lambda s: float(s.replace(",", "."))
                             if s and str(s).strip() not in ("", "None", "nan") else None
                            )(str(m.get("_volumen", "") or ""))
                        )
                        for m in st.session_state.get(_acomp_key, [])
                    ]
                    _acomp_vol_muestras = int(sum(v for v in _acomp_vols_f if v is not None))
                    _acomp_pond_st, _acomp_pond_ic = [], []
                    for _am in st.session_state.get(_acomp_key, []):
                        def _pn_form(x):
                            try: return float(str(x).replace(",", "."))
                            except: return None
                        _av  = _pn_form(_am.get("_volumen"))
                        _ast = _pn_form(_am.get("_st"))
                        _aic = _pn_form(_am.get("_ic"))
                        _acomp_pond_st.append(_av * _ast if _av is not None and _ast is not None else None)
                        _acomp_pond_ic.append(_av * _aic if _av is not None and _aic is not None else None)
                    _acomp_vol_total_f = sum(v for v in _acomp_vols_f if v is not None)
                    _acomp_st_pond = (round(sum(x for x in _acomp_pond_st if x is not None) / _acomp_vol_total_f, 2)
                                      if _acomp_vol_total_f else None)
                    _acomp_ic_pond = (round(sum(x for x in _acomp_pond_ic if x is not None) / _acomp_vol_total_f, 3)
                                      if _acomp_vol_total_f else None)
                    _acomp_diferencia_vol = (int(seg_vol_declarado) - _acomp_vol_muestras
                                             if seg_vol_declarado is not None else None)

                # ── CONTRAMUESTRAS: agregar muestra ──────────────────────
                if _sub == "CONTRAMUESTRAS SOLICITADAS":
                    _contra_key = "contra_muestras"
                    if _contra_key not in st.session_state:
                        st.session_state[_contra_key] = []
                    if st.button("➕  AGREGAR CONTRAMUESTRA", use_container_width=True,
                                 key=f"btn_agregar_contra_{_ti}"):
                        st.session_state[_contra_key].append({
                            "ID": seg_id_muestra or "",
                            "GRASA (%)": f"{seg_grasa:.2f}" if seg_grasa is not None else "",
                            "ST (%)":    f"{seg_st:.2f}"    if seg_st    is not None else "",
                            "IC (°C)":   f"{seg_ic:.3f}"    if seg_ic    is not None else "",
                            "AGUA (%)":  f"{seg_agua:.2f}"  if seg_agua  is not None else "",
                            "ALCOHOL":   seg_alcohol, "CLORUROS": seg_cloruros,
                            "NEUTRALIZANTES": seg_neutralizantes,
                            "OBS":       seg_observaciones or "",
                            "_grasa": seg_grasa, "_st": seg_st, "_ic": seg_ic,
                            "_proteina": seg_proteina,
                            "_agua": seg_agua, "_alcohol": seg_alcohol,
                            "_cloruros": seg_cloruros, "_neutralizantes": seg_neutralizantes,
                            "_obs": seg_observaciones or "",
                        })
                        st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                        st.rerun()
                    if st.session_state[_contra_key]:
                        st.markdown(
                            f"""<div style="font-size:0.9rem;font-weight:700;color:#0056A3;
                                           margin:10px 0 4px 0;">
                                 📋 {len(st.session_state[_contra_key])} contramuestra(s) registrada(s)
                               </div>""",
                            unsafe_allow_html=True,
                        )
                        df_prev_c = pd.DataFrame([
                            {k: v for k, v in m.items() if not k.startswith("_")}
                            for m in st.session_state[_contra_key]
                        ])
                        st.dataframe(df_prev_c, use_container_width=True, hide_index=True)

                # ── Imágenes de Muestras (SEGUIMIENTOS) ──────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         📷 Imágenes de Muestras
                       </div>""",
                    unsafe_allow_html=True,
                )
                _s_imgs_conf_key   = f"seg_imgs_confirmadas_{_ti}"
                _s_imgs_noms_key   = f"seg_imgs_nombres_{_ti}"
                _s_img_gen_key     = f"_seg_img_gen_{_ti}"
                if _s_imgs_conf_key not in st.session_state:
                    st.session_state[_s_imgs_conf_key] = False
                if _s_imgs_noms_key not in st.session_state:
                    st.session_state[_s_imgs_noms_key] = []
                _s_img_gen = st.session_state.get(_s_img_gen_key, 0)

                _s_imgs_subidas = st.file_uploader(
                    "ADJUNTAR IMÁGENES",
                    type=["png", "jpg", "jpeg"],
                    accept_multiple_files=True,
                    key=f"seg_imgs_uploader_{_ti}_{_s_img_gen}",
                    label_visibility="visible",
                )
                if _s_imgs_subidas:
                    _s_nombres = [f.name for f in _s_imgs_subidas]
                    if _s_nombres != st.session_state[_s_imgs_noms_key]:
                        st.session_state[_s_imgs_conf_key] = False
                    _s_confirmed = st.session_state[_s_imgs_conf_key]
                    _s_thumb = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
                    for _si in _s_imgs_subidas:
                        _sb64 = base64.b64encode(_si.read()).decode()
                        _sext = _si.name.rsplit(".", 1)[-1].lower()
                        _smime = "image/jpeg" if _sext in ("jpg", "jpeg") else "image/png"
                        _snc   = _si.name if len(_si.name) <= 16 else _si.name[:14] + "…"
                        _schk  = ("<div style='color:#16a34a;font-size:12px;"
                                  "text-align:center;font-weight:600;'>✅ Guardada</div>"
                                  if _s_confirmed else
                                  f"<div style='font-size:10px;color:#888;text-align:center;'>{_snc}</div>")
                        _sbrd  = "#16a34a" if _s_confirmed else "#D1D5DB"
                        _s_thumb += (f"<div style='display:flex;flex-direction:column;"
                                     f"align-items:center;gap:4px;'>"
                                     f"<img src='data:{_smime};base64,{_sb64}' "
                                     f"style='width:150px;height:150px;object-fit:cover;"
                                     f"border-radius:10px;border:2px solid {_sbrd};"
                                     f"box-shadow:0 2px 6px rgba(0,0,0,0.08);'/>"
                                     f"{_schk}</div>")
                        _si.seek(0)
                    _s_thumb += "</div>"
                    st.markdown(_s_thumb, unsafe_allow_html=True)
                    if not st.session_state[_s_imgs_conf_key]:
                        if st.button("💾 GUARDAR IMÁGENES",
                                     key=f"btn_seg_save_imgs_{_ti}",
                                     use_container_width=False):
                            st.session_state[_s_imgs_conf_key] = True
                            st.session_state[_s_imgs_noms_key] = _s_nombres
                            st.rerun()
                    else:
                        st.success("✅ Imágenes guardadas correctamente.")
                else:
                    st.session_state[_s_imgs_conf_key] = False
                    st.caption("No se han adjuntado imágenes.")

                # ── Guardar ───────────────────────────────────────────────
                st.markdown("---")
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         💾 Guardar en Historial
                       </div>""",
                    unsafe_allow_html=True,
                )
                if st.button(
                    f"💾  GUARDAR {_sub}", type="primary",
                    use_container_width=True, key=f"btn_guardar_seg_{_ti}",
                ):
                    ts = now_col().strftime("%d/%m/%Y %H:%M")
                    # ── Guardar fotos a disco ──────────────────────────────
                    _s_fotos_prefix = f"SEG_{_sub[:3].upper()}_{(seg_codigo or seg_quien_trajo or seg_responsable or 'X').replace(' ','_')[:12]}"
                    _s_fotos_paths  = save_fotos_to_disk(
                        _s_imgs_subidas or [], _s_fotos_prefix
                    )
                    base = {
                        "tipo_seguimiento":     "SEGUIMIENTOS",
                        "sub_tipo_seguimiento": _sub,
                        "fecha":                seg_fecha.strftime("%d/%m/%Y") if seg_fecha else "",
                        "seg_codigo":           seg_codigo,
                        "seg_quien_trajo":      seg_quien_trajo,
                        "ruta":                 seg_ruta_acomp,
                        "seg_responsable":      seg_responsable,
                        "guardado_en":          ts,
                        "fotos_json":           json.dumps(_s_fotos_paths, ensure_ascii=False),
                    }
                    def _guardar_lista_muestras(lista):
                        for m in lista:
                            save_seguimiento_to_csv({**base,
                                "seg_id_muestra":     m.get("ID", ""),
                                "seg_volumen":        m.get("_volumen", ""),
                                "seg_grasa":          m.get("_grasa", ""),
                                "seg_st":             m.get("_st", ""),
                                "seg_ic":             m.get("_ic", ""),
                                "seg_agua":           m.get("_agua", ""),
                                "seg_alcohol":        m.get("_alcohol", ""),
                                "seg_cloruros":       m.get("_cloruros", ""),
                                "seg_neutralizantes": m.get("_neutralizantes", ""),
                                "seg_observaciones":  m.get("_obs", ""),
                            })
                    if _sub == "ACOMPAÑAMIENTOS" and st.session_state.get("acomp_muestras"):
                        # Recalcular ponderados en el momento de guardar
                        _sv_muestras = st.session_state.acomp_muestras
                        def _pn_sv(x):
                            try: return float(str(x).replace(",", "."))
                            except: return None
                        _sv_vols = [_pn_sv(m.get("_volumen")) for m in _sv_muestras]
                        _sv_vol_tot = sum(v for v in _sv_vols if v is not None)
                        _sv_pond_st = [(_pn_sv(m.get("_volumen")) or 0) * (_pn_sv(m.get("_st")) or 0)
                                       for m in _sv_muestras
                                       if _pn_sv(m.get("_volumen")) is not None and _pn_sv(m.get("_st")) is not None]
                        _sv_pond_ic = [(_pn_sv(m.get("_volumen")) or 0) * (_pn_sv(m.get("_ic")) or 0)
                                       for m in _sv_muestras
                                       if _pn_sv(m.get("_volumen")) is not None and _pn_sv(m.get("_ic")) is not None]
                        _sv_st_pond = round(sum(_sv_pond_st) / _sv_vol_tot, 2) if _sv_vol_tot and _sv_pond_st else ""
                        _sv_ic_pond = round(sum(_sv_pond_ic) / _sv_vol_tot, 3) if _sv_vol_tot and _sv_pond_ic else ""
                        _sv_vol_dec = int(seg_vol_declarado) if seg_vol_declarado is not None else ""
                        _sv_vol_m   = int(_sv_vol_tot) if _sv_vol_tot else ""
                        _sv_dif_vol = (int(seg_vol_declarado) - int(_sv_vol_tot)
                                       if seg_vol_declarado is not None and _sv_vol_tot else "")
                        # Una sola fila con todas las muestras en JSON (igual que RUTAS)
                        save_seguimiento_to_csv({**base,
                            "seg_vol_declarado":   _sv_vol_dec,
                            "seg_vol_muestras":    _sv_vol_m,
                            "seg_diferencia_vol":  _sv_dif_vol,
                            "seg_solidos_ruta":    seg_solidos_ruta if seg_solidos_ruta is not None else "",
                            "seg_crioscopia_ruta": seg_crioscopia_ruta if seg_crioscopia_ruta is not None else "",
                            "seg_st_pond":         _sv_st_pond,
                            "seg_ic_pond":         _sv_ic_pond,
                            "muestras_json":       json.dumps(_sv_muestras, ensure_ascii=False),
                        })
                        st.session_state.acomp_muestras = []
                    elif _sub == "CONTRAMUESTRAS SOLICITADAS" and st.session_state.get("contra_muestras"):
                        _guardar_lista_muestras(st.session_state.contra_muestras)
                        st.session_state.contra_muestras = []
                    else:
                        save_seguimiento_to_csv({**base,
                            "seg_id_muestra":    seg_id_muestra or "",
                            "seg_grasa":         seg_grasa if seg_grasa is not None else "",
                            "seg_st":            seg_st    if seg_st    is not None else "",
                            "seg_ic":            seg_ic    if seg_ic    is not None else "",
                            "seg_agua":          seg_agua  if seg_agua  is not None else "",
                            "seg_alcohol":       seg_alcohol,
                            "seg_cloruros":      seg_cloruros,
                            "seg_neutralizantes": seg_neutralizantes,
                            "seg_observaciones": seg_observaciones or "",
                        })
                    for _k in [f"seg_fecha_{_ti}", f"seg_codigo_{_ti}",
                                f"seg_nombre_{_ti}",
                                f"seg_quien_trajo_{_ti}", f"seg_ruta_acomp_{_ti}",
                                f"seg_responsable_{_ti}",
                                f"seg_vol_declarado_{_ti}",
                                f"seg_solidos_ruta_{_ti}",
                                f"seg_crios_raw_{_ti}"]:
                        st.session_state.pop(_k, None)
                    st.session_state[_s_img_gen_key]  = _s_img_gen + 1
                    st.session_state[_s_imgs_conf_key] = False
                    st.session_state[_s_imgs_noms_key] = []
                    st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                    st.session_state[f"seg_guardado_ok_{_ti}"] = True
                    clear_draft_state()
                    st.rerun()

                if st.session_state.get(f"seg_guardado_ok_{_ti}"):
                    st.success(f"✅ Seguimiento {_sub} guardado en el historial.")
                    st.session_state[f"seg_guardado_ok_{_ti}"] = False


    if st.sidebar.button("REINICIAR FORMULARIO"):
        st.session_state.continuar = False
        clear_draft_state()
        st.rerun()



elif st.session_state.pagina_activa == "HISTORIAL":
    st.markdown("---")

    # ── HISTORIAL DE RUTAS ──────────────────────────────────────────────────────
    st.markdown(
        """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                       margin:14px 0 6px 0;letter-spacing:.4px;
                       border-left:4px solid #0056A3;padding-left:10px;">
             📊 Historial de Rutas
           </div>""",
        unsafe_allow_html=True,
    )

    df_hist = load_historial()
    if df_hist.empty:
        st.info("No hay rutas guardadas aún. Complete el formulario y presione **GUARDAR RUTA** para registrar datos aquí.")
    else:
        # ── Filtros ───────────────────────────────────────────────────
        st.markdown(
            "<div style='font-weight:600;color:#374151;margin-bottom:8px;'>"
            "🔍 Filtros de búsqueda</div>",
            unsafe_allow_html=True,
        )
        hf1, hf2, hf3, hf4 = st.columns([2, 2, 2, 2])

        # 1. Tipo de Seguimiento
        with hf1:
            filtro_tipo = st.selectbox(
                "TIPO DE SEGUIMIENTO",
                ["TODOS", "RUTAS", "TRANSUIZA", "SEGUIMIENTOS"],
                key="hist_tipo",
            )

        # 2. Rango de fechas
        fechas_validas = (
            df_hist["_fecha_dt"].dropna()
            if "_fecha_dt" in df_hist.columns
            else pd.Series(dtype="datetime64[ns]")
        )
        _hoy_col = now_col().date()
        fecha_min_val = _hoy_col
        fecha_max_val = _hoy_col

        with hf2:
            fecha_desde = st.date_input(
                "FECHA DESDE", value=fecha_min_val,
                format="DD/MM/YYYY", key="hist_desde",
            )
        with hf3:
            fecha_hasta = st.date_input(
                "FECHA HASTA", value=fecha_max_val,
                format="DD/MM/YYYY", key="hist_hasta",
            )

        # 3. Placa del Vehículo
        placas_unicas = (
            ["TODAS"] + sorted(df_hist["placa"].dropna().replace("", pd.NA).dropna().unique().tolist())
            if "placa" in df_hist.columns else ["TODAS"]
        )
        with hf4:
            filtro_placa = st.selectbox("PLACA DEL VEHÍCULO", placas_unicas, key="hist_placa")

        # Filtro adicional: Nombre de Ruta (solo visible cuando tipo = RUTAS)
        filtro_ruta = "TODAS"
        if filtro_tipo == "RUTAS":
            df_rutas_solo = df_hist[df_hist["tipo_seguimiento"] == "RUTAS"] \
                if "tipo_seguimiento" in df_hist.columns else df_hist
            rutas_unicas = (
                ["TODAS"] + sorted(
                    df_rutas_solo["ruta"].dropna().replace("", pd.NA).dropna().unique().tolist()
                )
                if "ruta" in df_rutas_solo.columns else ["TODAS"]
            )
            _rc, _rb, _ = st.columns([2, 1, 3])
            with _rc:
                filtro_ruta = st.selectbox(
                    "📍 NOMBRE DE RUTA", rutas_unicas, key="hist_ruta"
                )
            with _rb:
                st.markdown("<div style='margin-top:26px;'></div>", unsafe_allow_html=True)
                if st.button("🔍 BUSCAR", type="primary",
                             key="btn_buscar_rutas", use_container_width=True):
                    st.session_state.hist_buscar_ok = True
                    st.rerun()
        else:
            _rb2, _ = st.columns([1, 5])
            with _rb2:
                st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)
                if st.button("🔍 BUSCAR", type="primary",
                             key="btn_buscar_hist", use_container_width=True):
                    st.session_state.hist_buscar_ok = True
                    st.rerun()

        # ── Aplicar filtros ───────────────────────────────────────────
        if not st.session_state.hist_buscar_ok and not st.session_state.get("admin_accion"):
            st.info("Selecciona los filtros y presiona **🔍 BUSCAR** para ver el historial.")
        else:
            if filtro_tipo == "SEGUIMIENTOS":
                # Fuente separada: seguimientos_historial.csv
                df_filtrado = load_seguimientos()
                if "_fecha_dt" in df_filtrado.columns:
                    df_filtrado = df_filtrado[
                        (df_filtrado["_fecha_dt"].dt.date >= fecha_desde) &
                        (df_filtrado["_fecha_dt"].dt.date <= fecha_hasta)
                    ]
            else:
                df_filtrado = df_hist.copy()
                if "_fecha_dt" in df_filtrado.columns:
                    df_filtrado = df_filtrado[
                        (df_filtrado["_fecha_dt"].dt.date >= fecha_desde) &
                        (df_filtrado["_fecha_dt"].dt.date <= fecha_hasta)
                    ]
                if filtro_tipo != "TODOS" and "tipo_seguimiento" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["tipo_seguimiento"] == filtro_tipo]
                if filtro_placa != "TODAS" and "placa" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["placa"] == filtro_placa]
                if filtro_ruta != "TODAS" and "ruta" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["ruta"] == filtro_ruta]

            # ── Columna Estado de Calidad ─────────────────────────────────
            df_filtrado = df_filtrado.copy()
            df_filtrado["_estado"] = df_filtrado.apply(
                lambda r: calcular_estado_calidad(r.to_dict()), axis=1
            )

            # ── Tabla con filas en rojo para desviaciones ─────────────────
            n_desv = (df_filtrado["_estado"] == "DESVIACIÓN").sum()
            col_info, col_alerta = st.columns([3, 1])
            with col_info:
                st.markdown(f"**{len(df_filtrado)} registro(s) encontrado(s)**")
            with col_alerta:
                if n_desv:
                    st.markdown(
                        f"<div style='text-align:right;font-size:13px;"
                        f"color:#9C0006;font-weight:600;'>"
                        f"⚠️ {n_desv} desviación(es)</div>",
                        unsafe_allow_html=True,
                    )

            RED = "background-color:#FFC7CE;color:#9C0006;font-weight:700"

            # ── Columnas y styler según tipo de seguimiento ───────────────
            if filtro_tipo == "TRANSUIZA":
                # Columnas exclusivas TRANSUIZA
                cols_sel = ["tipo_seguimiento", "fecha", "placa",
                            "st_carrotanque", "grasa_muestra", "solidos_ruta",
                            "proteina_muestra", "diferencia_solidos", "guardado_en"]
                col_labels = {
                    "tipo_seguimiento": "TIPO", "fecha": "FECHA", "placa": "PLACA",
                    "st_carrotanque": "ST CARROTANQUE (%)",
                    "grasa_muestra": "GRASA (%)", "solidos_ruta": "ST MUESTRA (%)",
                    "proteina_muestra": "PROTEÍNA (%)",
                    "diferencia_solidos": "DIF. SÓLIDOS", "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "ST CARROTANQUE (%)": st.column_config.NumberColumn(format="%.2f"),
                    "GRASA (%)":          st.column_config.NumberColumn(format="%.2f"),
                    "ST MUESTRA (%)":     st.column_config.NumberColumn(format="%.2f"),
                    "PROTEÍNA (%)":       st.column_config.NumberColumn(format="%.2f"),
                    "DIF. SÓLIDOS":       st.column_config.NumberColumn(format="%.2f"),
                }
                def resaltar_celdas(row):
                    return [""] * len(row)

            elif filtro_tipo == "RUTAS":
                # Columnas completas RUTAS
                cols_sel = ["tipo_seguimiento", "fecha", "ruta", "placa", "conductor",
                            "volumen_declarado", "vol_estaciones", "diferencia",
                            "solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
                            "num_estaciones", "guardado_en"]
                col_labels = {
                    "tipo_seguimiento": "TIPO", "fecha": "FECHA", "ruta": "RUTA",
                    "placa": "PLACA", "conductor": "CONDUCTOR",
                    "volumen_declarado": "VOL. DECL. (L)", "vol_estaciones": "VOL. EST. (L)",
                    "diferencia": "DIFER. (L)", "solidos_ruta": "ST RUTA (%)",
                    "crioscopia_ruta": "IC RUTA (°C)", "st_pond": "ST POND",
                    "ic_pond": "IC POND", "num_estaciones": "Nº EST.",
                    "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "ST RUTA (%)":      st.column_config.NumberColumn(format="%.2f"),
                    "ST POND":          st.column_config.NumberColumn(format="%.2f"),
                    "IC RUTA (°C)":     st.column_config.NumberColumn(format="%.3f"),
                    "IC POND":          st.column_config.NumberColumn(format="%.3f"),
                    "VOL. DECL. (L)":   st.column_config.NumberColumn(format="%d"),
                    "VOL. EST. (L)":    st.column_config.NumberColumn(format="%d"),
                    "DIFER. (L)":       st.column_config.NumberColumn(format="%d"),
                    "Nº EST.":          st.column_config.NumberColumn(format="%d"),
                }
                def resaltar_celdas(row):
                    styles = [""] * len(row)
                    cols = list(row.index)
                    desv_st = desv_ic = False
                    try:
                        v = float(str(row.get("ST RUTA (%)", "")).replace(",", "."))
                        if 0 < v < 12.60: desv_st = True
                    except Exception: pass
                    try:
                        v = float(str(row.get("IC RUTA (°C)", "")).replace(",", "."))
                        if v > -0.535 or v < -0.550: desv_ic = True
                    except Exception: pass
                    if (desv_st or desv_ic) and "RUTA" in cols:
                        styles[cols.index("RUTA")] = RED
                    if desv_st and "ST RUTA (%)" in cols:
                        styles[cols.index("ST RUTA (%)")] = RED
                    if desv_ic and "IC RUTA (°C)" in cols:
                        styles[cols.index("IC RUTA (°C)")] = RED
                    return styles

            elif filtro_tipo == "SEGUIMIENTOS":
                cols_sel = ["sub_tipo_seguimiento", "fecha", "seg_codigo",
                            "seg_quien_trajo", "ruta", "seg_responsable",
                            "seg_id_muestra", "seg_grasa", "seg_st", "seg_ic", "seg_agua",
                            "seg_alcohol", "seg_cloruros", "seg_neutralizantes",
                            "seg_observaciones", "guardado_en"]
                col_labels = {
                    "sub_tipo_seguimiento": "SUB-TIPO", "fecha": "FECHA",
                    "seg_codigo": "CÓDIGO", "seg_quien_trajo": "ENTREGADO POR",
                    "ruta": "RUTA", "seg_responsable": "RESPONSABLE",
                    "seg_id_muestra": "ID MUESTRA",
                    "seg_grasa": "GRASA (%)", "seg_st": "ST (%)",
                    "seg_ic": "IC (°C)", "seg_agua": "AGUA (%)",
                    "seg_alcohol": "ALCOHOL", "seg_cloruros": "CLORUROS",
                    "seg_neutralizantes": "NEUTRALIZANTES",
                    "seg_observaciones": "OBSERVACIONES", "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "GRASA (%)": st.column_config.NumberColumn(format="%.2f"),
                    "ST (%)":    st.column_config.NumberColumn(format="%.2f"),
                    "IC (°C)":   st.column_config.NumberColumn(format="%.3f"),
                    "AGUA (%)":  st.column_config.NumberColumn(format="%.2f"),
                }
                def resaltar_celdas(row):
                    return [""] * len(row)

            else:
                # TODOS: columnas comunes + identificación
                cols_sel = ["tipo_seguimiento", "fecha", "ruta", "placa", "conductor",
                            "solidos_ruta", "crioscopia_ruta", "guardado_en"]
                col_labels = {
                    "tipo_seguimiento": "TIPO", "fecha": "FECHA", "ruta": "RUTA",
                    "placa": "PLACA", "conductor": "CONDUCTOR",
                    "solidos_ruta": "ST / ST MUESTRA (%)", "crioscopia_ruta": "IC RUTA (°C)",
                    "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "ST / ST MUESTRA (%)": st.column_config.NumberColumn(format="%.2f"),
                    "IC RUTA (°C)":        st.column_config.NumberColumn(format="%.3f"),
                }
                def resaltar_celdas(row):
                    styles = [""] * len(row)
                    # Solo resaltar filas de RUTAS, no TRANSUIZA ni SEGUIMIENTOS
                    if str(row.get("TIPO", "")).strip() == "TRANSUIZA":
                        return styles
                    cols = list(row.index)
                    desv_st = desv_ic = False
                    try:
                        v = float(str(row.get("ST / ST MUESTRA (%)", "")).replace(",", "."))
                        if 0 < v < 12.60: desv_st = True
                    except Exception: pass
                    try:
                        v = float(str(row.get("IC RUTA (°C)", "")).replace(",", "."))
                        if v > -0.535 or v < -0.550: desv_ic = True
                    except Exception: pass
                    if (desv_st or desv_ic) and "RUTA" in cols:
                        styles[cols.index("RUTA")] = RED
                    if desv_st and "ST / ST MUESTRA (%)" in cols:
                        styles[cols.index("ST / ST MUESTRA (%)")] = RED
                    if desv_ic and "IC RUTA (°C)" in cols:
                        styles[cols.index("IC RUTA (°C)")] = RED
                    return styles

            # Filtrar solo columnas que existen
            cols_data = [c for c in CSV_COLS if c in df_filtrado.columns]
            cols_vis   = [c for c in cols_sel if c in df_filtrado.columns]
            df_display = df_filtrado[cols_vis].rename(columns=col_labels).reset_index(drop=True)

            sel = st.dataframe(
                df_display.style.apply(resaltar_celdas, axis=1),
                use_container_width=True,
                hide_index=True,
                on_select="rerun",
                selection_mode="multi-row",
                column_config=col_config_map,
            )

            # ── Descarga Excel según filtros activos ──────────────────────
            if not df_filtrado.empty:
                _ts = now_col().strftime('%Y%m%d_%H%M')
                _cx, _ = st.columns([1, 3])
                with _cx:
                    st.download_button(
                        label="⬇️ DESCARGAR REPORTE EXCEL",
                        data=historial_to_excel_filtrado(
                            df_filtrado, fecha_desde, fecha_hasta, filtro_tipo
                        ),
                        file_name=f"historial_qualilact_{_ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )

            # ── Botones de acción — visibles al seleccionar una o varias filas ─
            orig_indices = df_filtrado.index.tolist()
            filas_sel    = (sel.selection.rows
                            if sel and hasattr(sel, "selection") else [])
            sel_orig_idxs = [orig_indices[i] for i in filas_sel if i < len(orig_indices)]
            sel_orig_idx  = sel_orig_idxs[0] if len(sel_orig_idxs) == 1 else None

            if sel_orig_idxs:
                n_sel = len(sel_orig_idxs)
                st.markdown(
                    f"<div style='font-size:12px;color:#6B7280;margin:6px 0 4px 0;'>"
                    f"{'1 fila seleccionada' if n_sel == 1 else f'{n_sel} filas seleccionadas'}"
                    f" — elige una acción:</div>",
                    unsafe_allow_html=True,
                )
                ab1, ab2, _ = st.columns([1, 1, 5])
                with ab1:
                    # Modificar solo disponible con exactamente 1 fila
                    if st.button("✏️ Modificar", key="btn_modificar",
                                 use_container_width=True,
                                 help="Editar este registro",
                                 disabled=(n_sel != 1)):
                        st.session_state.admin_accion = "modificar"
                        st.session_state.admin_idx    = sel_orig_idx
                        st.rerun()
                with ab2:
                    if st.button("🗑️ Eliminar", key="btn_eliminar",
                                 use_container_width=True,
                                 help=f"Eliminar {n_sel} registro(s) seleccionado(s)"):
                        st.session_state.admin_accion = "eliminar"
                        st.session_state.admin_idxs   = sel_orig_idxs
                        st.session_state.admin_idx    = sel_orig_idx
                        st.rerun()

            # ── Detalle de ruta al seleccionar una fila ────────────────────
            # Si estamos en modo edición y la selección se perdió en el rerun,
            # recuperar el índice desde la sesión para mantener el panel abierto.
            _ss_accion = st.session_state.get("admin_accion")
            _ss_idx    = st.session_state.get("admin_idx")
            if (sel_orig_idx is None and _ss_accion == "modificar"
                    and _ss_idx is not None):
                sel_orig_idx = _ss_idx

            # ── Panel de detalle ACOMPAÑAMIENTOS ───────────────────────────
            if (sel_orig_idx is not None and filtro_tipo == "SEGUIMIENTOS"):
                if sel_orig_idx in df_filtrado.index:
                    _srow = df_filtrado.loc[sel_orig_idx]
                else:
                    _srow = {}
                _s_sub = str(_srow.get("sub_tipo_seguimiento", "")).strip()
                if _s_sub == "ACOMPAÑAMIENTOS":
                    _mj_raw = str(_srow.get("muestras_json", "") or "").strip()
                    try:
                        _mj_data = json.loads(_mj_raw) if _mj_raw else []
                    except Exception:
                        _mj_data = []
                    if _mj_data:
                        st.markdown("---")
                        _ac_entreg  = str(_srow.get("seg_quien_trajo", "") or "").strip() or "—"
                        _ac_ruta    = str(_srow.get("ruta", "") or "").strip() or "—"
                        _ac_resp    = str(_srow.get("seg_responsable", "") or "").strip() or "—"
                        _ac_cod     = str(_srow.get("seg_codigo", "") or "").strip() or "—"
                        _ac_fecha   = str(_srow.get("fecha", "") or "").strip() or "—"
                        st.markdown(
                            f"""<div style="background:#0056A3;border-radius:8px;
                                           padding:10px 16px;margin-bottom:14px;">
                                  <span style="font-size:1rem;font-weight:700;color:#fff;
                                               letter-spacing:.05em;">
                                    👥 DETALLE ACOMPAÑAMIENTO
                                  </span>
                                  <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                                    {_ac_cod} &nbsp;·&nbsp; {_ac_fecha}
                                  </span>
                                </div>""",
                            unsafe_allow_html=True,
                        )

                        # ── Helper KPI (misma firma que el de RUTAS) ─────────────
                        def _kpi_card_ac(label, value, badge=None, badge_ok=True):
                            badge_html = ""
                            if badge:
                                bg  = "#D4EDDA" if badge_ok else "#F8D7DA"
                                col = "#155724" if badge_ok else "#721C24"
                                badge_html = (
                                    f'<div style="margin-top:4px;font-size:.65rem;font-weight:700;'
                                    f'color:{col};background:{bg};border-radius:4px;'
                                    f'padding:1px 5px;display:inline-block;">{badge}</div>'
                                )
                            return (
                                f'<div style="background:#fff;border:1px solid #dde6f0;'
                                f'border-radius:8px;padding:10px 12px;text-align:center;height:100%;">'
                                f'<div style="font-size:.62rem;font-weight:700;color:#6c8ca8;'
                                f'letter-spacing:.06em;margin-bottom:4px;">{label}</div>'
                                f'<div style="font-size:1.05rem;font-weight:800;color:#0056A3;">{value}</div>'
                                f'{badge_html}</div>'
                            )

                        # ── Cómputo de totales ────────────────────────────────────
                        def _pnac(x):
                            try: return float(str(x).replace(",", "."))
                            except: return None

                        _ac_n_muestras  = len(_mj_data)
                        _entreg_short   = (_ac_entreg[:16]+"…") if len(_ac_entreg) > 16 else _ac_entreg
                        _resp_short     = (_ac_resp[:16]+"…")   if len(_ac_resp)   > 16 else _ac_resp

                        # Leer campos guardados del CSV
                        _ac_vol_dec = _srow.get("seg_vol_declarado", "")
                        _ac_vol_m   = _srow.get("seg_vol_muestras",  "")
                        _ac_dif_vol = _srow.get("seg_diferencia_vol","")
                        _ac_st_r    = _srow.get("seg_solidos_ruta",  "")
                        _ac_ic_r    = _srow.get("seg_crioscopia_ruta","")
                        _ac_st_p    = _srow.get("seg_st_pond",       "")
                        _ac_ic_p    = _srow.get("seg_ic_pond",       "")

                        try:    _v_ac_vol_dec = f"{int(float(_ac_vol_dec)):,} L"
                        except: _v_ac_vol_dec = "—"
                        try:    _v_ac_vol_m   = f"{int(float(_ac_vol_m)):,} L"
                        except: _v_ac_vol_m   = "—"
                        try:
                            _dif_v = int(float(_ac_dif_vol))
                            _dif_ok = abs(_dif_v) <= 20
                            _v_ac_dif = f"{_dif_v:+,} L"
                            _b_ac_dif = ("✔ OK" if _dif_ok else "⚠ DIFERENCIA", _dif_ok)
                        except: _v_ac_dif = "—"; _b_ac_dif = (None, True)
                        try:
                            _st_rv = float(_ac_st_r)
                            _st_ok = _st_rv >= 12.60
                            _v_ac_st_r = f"{_st_rv:.2f} %"
                            _b_ac_st   = ("✔ CONFORME" if _st_ok else "✖ DESVIACIÓN", _st_ok)
                        except: _v_ac_st_r = "—"; _b_ac_st = (None, True)
                        try:
                            _ic_rv = float(_ac_ic_r)
                            _ic_ok = -0.550 <= _ic_rv <= -0.535
                            _v_ac_ic_r = f"{_ic_rv:.3f} °C"
                            _b_ac_ic   = ("✔ CONFORME" if _ic_ok else "✖ DESVIACIÓN", _ic_ok)
                        except: _v_ac_ic_r = "—"; _b_ac_ic = (None, True)
                        try:    _v_ac_st_p = f"{float(_ac_st_p):.2f} %"
                        except: _v_ac_st_p = "—"
                        try:    _v_ac_ic_p = f"{float(_ac_ic_p):.3f} °C"
                        except: _v_ac_ic_p = "—"
                        try:    _v_ac_dif_st = f"{float(_ac_st_r) - float(_ac_st_p):+.2f} %"
                        except: _v_ac_dif_st = "—"
                        try:    _v_ac_dif_ic = f"{float(_ac_ic_r) - float(_ac_ic_p):+.3f} °C"
                        except: _v_ac_dif_ic = "—"

                        # ── Fila 1: identificación ─────────────────────────────────
                        _ac_row1_html = (
                            '<div style="display:grid;grid-template-columns:repeat(3,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_card_ac("ENTREGADO POR",  _entreg_short)
                            + _kpi_card_ac("RUTA",           _ac_ruta)
                            + _kpi_card_ac("Nº MUESTRAS",    str(_ac_n_muestras))
                            + '</div>'
                        )
                        # ── Fila 2: volúmenes (igual que RUTAS) ────────────────────
                        _ac_row2_html = (
                            '<div style="display:grid;grid-template-columns:repeat(3,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_card_ac("VOL. DECLARADO",    _v_ac_vol_dec)
                            + _kpi_card_ac("VOL. SUMA MUESTRAS", _v_ac_vol_m)
                            + _kpi_card_ac("DIFERENCIA VOL.",   _v_ac_dif, _b_ac_dif[0], _b_ac_dif[1])
                            + '</div>'
                        )
                        # ── Fila 3: ST e IC con ponderados y deltas ────────────────
                        _ac_row3_html = (
                            '<div style="display:grid;grid-template-columns:repeat(6,1fr);'
                            'gap:8px;margin-bottom:14px;">'
                            + _kpi_card_ac("ST RUTA (%)",      _v_ac_st_r, _b_ac_st[0], _b_ac_st[1])
                            + _kpi_card_ac("ST PONDERADO",     _v_ac_st_p)
                            + _kpi_card_ac("ΔST (RUTA−POND)",  _v_ac_dif_st)
                            + _kpi_card_ac("IC RUTA (°C)",     _v_ac_ic_r, _b_ac_ic[0], _b_ac_ic[1])
                            + _kpi_card_ac("IC PONDERADO",     _v_ac_ic_p)
                            + _kpi_card_ac("ΔIC (RUTA−POND)",  _v_ac_dif_ic)
                            + '</div>'
                        )
                        st.markdown(_ac_row1_html + _ac_row2_html + _ac_row3_html, unsafe_allow_html=True)

                        st.markdown(
                            "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                            "letter-spacing:.05em;margin-bottom:4px;'>"
                            "📋 CALIDAD POR MUESTRA</div>",
                            unsafe_allow_html=True,
                        )
                        _cat_ac = load_catalogo()
                        _cat_ac_map = dict(zip(_cat_ac["codigo"], _cat_ac["nombre"]))
                        _acomp_det_rows = []
                        for _am in _mj_data:
                            _cod_am = str(_am.get("ID", "") or "").strip()
                            _vol_am = _pnac(_am.get("_volumen"))
                            _st_am  = _pnac(_am.get("_st"))
                            _ic_am  = _pnac(_am.get("_ic"))
                            _pst_am = round(_vol_am * _st_am, 2) if _vol_am is not None and _st_am is not None else None
                            _pic_am = round(_vol_am * _ic_am, 3) if _vol_am is not None and _ic_am is not None else None
                            _acomp_det_rows.append({
                                "CÓDIGO":         _cod_am,
                                "NOMBRE ESTACIÓN": _cat_ac_map.get(_cod_am, ""),
                                "GRASA (%)":      _pnac(_am.get("_grasa")),
                                "ST (%)":         _st_am,
                                "PROTEÍNA (%)":   _pnac(_am.get("_proteina")),
                                "IC (°C)":        _ic_am,
                                "AGUA (%)":       _pnac(_am.get("_agua")),
                                "VOLUMEN (L)":    int(_vol_am) if _vol_am is not None else None,
                                "POND ST":        _pst_am,
                                "IC POND":        _pic_am,
                                "ALCOHOL":        _am.get("_alcohol", "N/A") or "N/A",
                                "CLORUROS":       _am.get("_cloruros", "N/A") or "N/A",
                                "NEUTRALIZANTES": _am.get("_neutralizantes", "N/A") or "N/A",
                                "OBS":            _am.get("_obs", "") or "",
                            })
                        _df_ac_det = pd.DataFrame(_acomp_det_rows)
                        _RED_AC = "background-color:#FFC7CE;color:#9C0006;font-weight:700"
                        def _color_ac(row):
                            styles = [""] * len(row)
                            cols = list(row.index)
                            try:
                                _ic_v = row.get("IC (°C)")
                                if _ic_v is not None:
                                    _icf = float(_ic_v)
                                    if (_icf > -0.530 or _icf < -0.550) and "IC (°C)" in cols:
                                        styles[cols.index("IC (°C)")] = _RED_AC
                            except Exception: pass
                            for _qc in ("ALCOHOL", "CLORUROS", "NEUTRALIZANTES"):
                                try:
                                    if row.get(_qc) == "+" and _qc in cols:
                                        styles[cols.index(_qc)] = _RED_AC
                                except Exception: pass
                            return styles
                        _fmt_ac = {
                            "GRASA (%)": "{:.2f}", "ST (%)": "{:.2f}",
                            "PROTEÍNA (%)": "{:.2f}",
                            "IC (°C)": "{:.3f}",   "AGUA (%)": "{:.2f}",
                            "POND ST": "{:.2f}",   "IC POND": "{:.3f}",
                        }
                        st.dataframe(
                            _df_ac_det.style.apply(_color_ac, axis=1).format(_fmt_ac, na_rep="—"),
                            use_container_width=True, hide_index=True,
                            height=min(38 + 35 * len(_acomp_det_rows), 420),
                            column_config={
                                "CÓDIGO":           st.column_config.TextColumn("CÓDIGO",          width="small"),
                                "NOMBRE ESTACIÓN":  st.column_config.TextColumn("NOMBRE ESTACIÓN", width="medium"),
                                "GRASA (%)":        st.column_config.NumberColumn("GRASA (%)",      width="small", format="%.2f"),
                                "ST (%)":           st.column_config.NumberColumn("ST (%)",         width="small", format="%.2f"),
                                "PROTEÍNA (%)":     st.column_config.NumberColumn("PROTEÍNA (%)",   width="small", format="%.2f"),
                                "IC (°C)":          st.column_config.NumberColumn("IC (°C)",        width="small", format="%.3f"),
                                "AGUA (%)":         st.column_config.NumberColumn("AGUA (%)",       width="small", format="%.2f"),
                                "VOLUMEN (L)":      st.column_config.NumberColumn("VOL. (L)",       width="small", format="%d"),
                                "POND ST":          st.column_config.NumberColumn("POND ST",        width="small", format="%.2f"),
                                "IC POND":          st.column_config.NumberColumn("IC POND",        width="small", format="%.3f"),
                                "ALCOHOL":          st.column_config.TextColumn("ALC.",             width="small"),
                                "CLORUROS":         st.column_config.TextColumn("CLOR.",            width="small"),
                                "NEUTRALIZANTES":   st.column_config.TextColumn("NEUT.",            width="small"),
                                "OBS":              st.column_config.TextColumn("OBSERVACIONES",    width="medium"),
                            },
                        )
                        # Fotos asociadas
                        _s_fotos_raw = str(_srow.get("fotos_json", "") or "").strip()
                        if _s_fotos_raw and _s_fotos_raw not in ("[]", ""):
                            try:
                                _s_fotos_list = json.loads(_s_fotos_raw)
                            except Exception:
                                _s_fotos_list = []
                            _s_fotos_ok = [p for p in _s_fotos_list if os.path.exists(p)]
                            if _s_fotos_ok:
                                st.markdown(
                                    "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                    "letter-spacing:.05em;margin:10px 0 6px;'>"
                                    "📷 IMÁGENES DE MUESTRAS</div>",
                                    unsafe_allow_html=True,
                                )
                                _cols_sfoto = st.columns(min(len(_s_fotos_ok), 4))
                                for _sfi, _sfp in enumerate(_s_fotos_ok):
                                    with _cols_sfoto[_sfi % 4]:
                                        st.image(_sfp, use_container_width=True)

            if sel_orig_idx is not None and filtro_tipo in ("RUTAS", "TODOS", "TRANSUIZA"):
                if sel_orig_idx in df_filtrado.index:
                    _drow = df_filtrado.loc[sel_orig_idx]
                else:
                    _df_all = load_historial()
                    _drow   = (_df_all.loc[sel_orig_idx]
                               if sel_orig_idx in _df_all.index else {})
                _tipo_reg = str(_drow.get("tipo_seguimiento", "RUTAS")).strip()
                _d_nombre = str(_drow.get("ruta", "")).strip() or "—"
                _d_fecha  = str(_drow.get("fecha", "")).strip() or "—"
                _d_placa  = str(_drow.get("placa",     "")).strip() or "—"
                _d_cond   = str(_drow.get("conductor", "")).strip() or "—"
                _d_vold   = _drow.get("volumen_declarado", "")
                _d_vole   = _drow.get("vol_estaciones",    "")
                _d_dif    = _drow.get("diferencia",        "")
                _d_st     = _drow.get("solidos_ruta",      "")
                _d_ic     = _drow.get("crioscopia_ruta",   "")
                _d_stpond = _drow.get("st_pond",           "")
                _d_icpond = _drow.get("ic_pond",           "")
                _d_nest   = _drow.get("num_estaciones",    "")
                # Campos exclusivos TRANSUIZA
                _d_st_car = _drow.get("st_carrotanque",   "")
                _d_grasa  = _drow.get("grasa_muestra",    "")
                _d_prot   = _drow.get("proteina_muestra", "")
                _d_dif_s  = _drow.get("diferencia_solidos","")

                st.markdown("---")

                # ── Encabezado del panel ──────────────────────────────────
                _panel_titulo = "🔍 DETALLE TRANSUIZA" if _tipo_reg == "TRANSUIZA" else "🔍 DETALLE DE RUTA"
                st.markdown(
                    f"""<div style="background:#0056A3;border-radius:8px;
                                   padding:10px 16px;margin-bottom:14px;">
                          <span style="font-size:1rem;font-weight:700;color:#fff;
                                       letter-spacing:.05em;">
                            {_panel_titulo}
                          </span>
                          <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                            {_d_placa} &nbsp;·&nbsp; {_d_fecha}
                          </span>
                        </div>""",
                    unsafe_allow_html=True,
                )

                # ── Detección de modo edición ─────────────────────────────
                _det_accion = st.session_state.get("admin_accion")
                _det_idx    = st.session_state.get("admin_idx")
                _edit_mode  = (_det_accion == "modificar" and _det_idx == sel_orig_idx)

                if _edit_mode:
                    # ══ MODO EDICIÓN: ramas por tipo de registro ════
                    try:
                        _fe_orig = datetime.strptime(str(_drow.get("fecha", "")), "%d/%m/%Y").date()
                    except Exception:
                        _fe_orig = date.today()

                    if _tipo_reg == "TRANSUIZA":
                        # ── Edición TRANSUIZA ─────────────────────────────
                        try:
                            _stcar_orig = float(str(_d_st_car or 0).replace(",","."))
                        except Exception: _stcar_orig = 0.0
                        try:
                            _grasa_orig = float(str(_d_grasa or 0).replace(",","."))
                        except Exception: _grasa_orig = 0.0
                        try:
                            _stm_orig = float(str(_d_st or 0).replace(",","."))
                        except Exception: _stm_orig = 0.0
                        try:
                            _prot_orig = float(str(_d_prot or 0).replace(",","."))
                        except Exception: _prot_orig = 0.0

                        te1, te2 = st.columns(2)
                        with te1:
                            _te_fecha = st.date_input("FECHA", value=_fe_orig,
                                                      format="DD/MM/YYYY", key="edit_t_fecha")
                            _te_placa = st.text_input("PLACA", value=_d_placa, key="edit_t_placa")
                            _te_stcar = st.number_input("ST DEL CARROTANQUE (%)",
                                                        value=_stcar_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_stcar")
                        with te2:
                            _te_grasa = st.number_input("GRASA (%)",
                                                        value=_grasa_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_grasa")
                            _te_stm   = st.number_input("ST MUESTRA (%)",
                                                        value=_stm_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_stm")
                            _te_prot  = st.number_input("PROTEÍNA (%)",
                                                        value=_prot_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_prot")
                        _te_dif = round(_te_stcar - _te_stm, 2)
                        _dif_col = "#9C0006" if abs(_te_dif) > 0.5 else "#006100"
                        st.markdown(
                            f"<div style='text-align:center;padding:8px;background:#F8FAFC;"
                            f"border-radius:8px;border:1px solid #D1D5DB;'>"
                            f"<div style='font-size:11px;font-weight:600;color:#6B7280;'>"
                            f"DIFERENCIA DE SÓLIDOS</div>"
                            f"<div style='font-size:1.5rem;font-weight:800;color:{_dif_col};'>"
                            f"{_te_dif:+.2f} %</div></div>",
                            unsafe_allow_html=True,
                        )
                        tec1, tec2, _ = st.columns([1.5, 1, 3])
                        with tec1:
                            if st.button("💾 GUARDAR CAMBIOS", type="primary",
                                         key="btn_save_edit_t", use_container_width=True):
                                update_row_in_csv(_det_idx, {
                                    "fecha":              _te_fecha.strftime("%d/%m/%Y"),
                                    "placa":              str(_te_placa).upper(),
                                    "st_carrotanque":     round(float(_te_stcar), 2),
                                    "solidos_ruta":       round(float(_te_stm), 2),
                                    "grasa_muestra":      round(float(_te_grasa), 2),
                                    "proteina_muestra":   round(float(_te_prot), 2),
                                    "diferencia_solidos": _te_dif,
                                })
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with tec2:
                            if st.button("✖ CANCELAR", key="btn_cancel_edit_t",
                                         use_container_width=True):
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                    else:
                        # ── Edición RUTAS ─────────────────────────────────
                        try:
                            _vol_orig = int(float(str(_drow.get("volumen_declarado", 0) or 0)))
                        except Exception:
                            _vol_orig = 0
                        try:
                            _st_orig = float(str(_drow.get("solidos_ruta", "0") or 0).replace(",", "."))
                        except Exception:
                            _st_orig = 0.0
                        try:
                            _ic_orig = float(str(_drow.get("crioscopia_ruta", "0") or 0).replace(",", "."))
                        except Exception:
                            _ic_orig = 0.0

                        ef1, ef2 = st.columns(2)
                        with ef1:
                            _e_fecha = st.date_input("FECHA", value=_fe_orig,
                                                     format="DD/MM/YYYY", key="edit_fecha")
                            _e_ruta  = st.text_input("RUTA",
                                                     value=str(_drow.get("ruta", "")), key="edit_ruta")
                            _e_placa = st.text_input("PLACA",
                                                     value=str(_drow.get("placa", "")), key="edit_placa")
                            _e_cond  = st.text_input("CONDUCTOR",
                                                     value=str(_drow.get("conductor", "")), key="edit_cond")
                        with ef2:
                            _e_vol = st.number_input("VOLUMEN DECLARADO (L)",
                                                     value=_vol_orig, min_value=0, step=1, key="edit_vol")
                            _e_st  = st.number_input("ST RUTA (%)", value=_st_orig,
                                                     step=0.01, format="%.2f", key="edit_st")
                            _e_ic  = st.number_input("IC RUTA (°C)", value=_ic_orig,
                                                     step=0.001, format="%.3f", key="edit_ic")

                        st.markdown(
                            "<div style='font-weight:700;color:#0056A3;margin:12px 0 6px;"
                            "font-size:.9rem;border-left:4px solid #0056A3;padding-left:8px;'>"
                            "🏭 Estaciones</div>",
                            unsafe_allow_html=True,
                        )
                        _ECOLS_E   = ["codigo", "grasa", "solidos", "proteina",
                                       "crioscopia", "agua_pct", "volumen", "alcohol",
                                       "cloruros", "neutralizantes", "obs"]
                        _est_json_e = str(_drow.get("estaciones_json", "") or "").strip()
                        try:
                            _est_data_e = json.loads(_est_json_e) if _est_json_e else []
                        except Exception:
                            _est_data_e = []
                        # ── Caché de edición para preservar cambios entre reruns ──
                        _h_cache_key = f"_h_est_cache_{st.session_state.get('admin_idx', 0)}"
                        _cat_h = load_catalogo()
                        _cat_h_map = dict(zip(_cat_h["codigo"], _cat_h["nombre"]))

                        if _h_cache_key in st.session_state:
                            # Usar la versión cacheada (conserva ediciones previas)
                            _df_est_e = st.session_state[_h_cache_key].copy()
                        else:
                            _df_est_e = (pd.DataFrame(_est_data_e, columns=_ECOLS_E)
                                         if _est_data_e else pd.DataFrame(columns=_ECOLS_E))
                            for _nc in ["grasa", "solidos", "proteina", "agua_pct"]:
                                _df_est_e[_nc] = pd.to_numeric(_df_est_e[_nc], errors="coerce")
                            _df_est_e["volumen"] = pd.to_numeric(
                                _df_est_e["volumen"], errors="coerce")

                        # Siempre recomputar nombre desde el catálogo (refleja código actual)
                        _df_est_e["nombre_estacion"] = _df_est_e["codigo"].apply(
                            lambda c: _cat_h_map.get(str(c).strip(), "")
                                      if pd.notna(c) else ""
                        )

                        st.caption("💡 Tab → siguiente celda · Enter → siguiente fila")
                        _edited_df_e = st.data_editor(
                            _df_est_e, num_rows="dynamic", use_container_width=True,
                            key="edit_est_editor",
                            column_config={
                                "codigo":         st.column_config.TextColumn("CÓDIGO"),
                                "grasa":          st.column_config.NumberColumn("GRASA (%)",     format="%.2f", min_value=0.0, max_value=100.0),
                                "solidos":        st.column_config.NumberColumn("SÓL.TOT. (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                "proteina":       st.column_config.NumberColumn("PROTEÍNA (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                "crioscopia":     st.column_config.TextColumn("CRIOSCOPIA (°C)"),
                                "volumen":        st.column_config.NumberColumn("VOLUMEN (L)",   format="%.0f", min_value=0,   step=1),
                                "alcohol":        st.column_config.SelectboxColumn("ALCOHOL",        options=["N/A", "+", "-"], required=True),
                                "cloruros":       st.column_config.SelectboxColumn("CLORUROS",       options=["N/A", "+", "-"], required=True),
                                "neutralizantes": st.column_config.SelectboxColumn("NEUTRALIZANTES", options=["N/A", "+", "-"], required=True),
                                "agua_pct":       st.column_config.NumberColumn("% AGUA",        format="%.1f", min_value=0.0, max_value=100.0),
                                "obs":            st.column_config.TextColumn("OBSERVACIONES"),
                                "nombre_estacion": st.column_config.TextColumn(
                                                     "NOMBRE ESTACIÓN", disabled=True),
                            },
                            hide_index=True,
                        )

                        # Guardar edición en caché (sin nombre_estacion) y refrescar si cambió código
                        _h_prev_codes = list(_df_est_e["codigo"].fillna("").astype(str).str.strip())
                        _h_new_codes  = list(_edited_df_e["codigo"].fillna("").astype(str).str.strip())
                        _cache_df = _edited_df_e.drop(columns=["nombre_estacion"], errors="ignore").copy()
                        st.session_state[_h_cache_key] = _cache_df
                        if _h_prev_codes != _h_new_codes:
                            st.rerun()

                        # ── JS: Enter avanza a la siguiente celda ─────────────
                        st.components.v1.html("""
                        <script>
                        (function(){
                          function patch(){
                            document.querySelectorAll('[data-testid="stDataEditor"]').forEach(function(grid){
                              if(grid._ep) return;
                              grid._ep = true;
                              grid.addEventListener('keydown', function(e){
                                if(e.key!=='Enter') return;
                                var a = document.activeElement;
                                if(!a || (a.tagName!=='INPUT' && a.tagName!=='TEXTAREA')) return;
                                e.preventDefault(); e.stopPropagation();
                                a.dispatchEvent(new KeyboardEvent('keydown',{
                                  key:'Tab',code:'Tab',keyCode:9,which:9,
                                  bubbles:true,cancelable:true,composed:true
                                }));
                              }, true);
                            });
                          }
                          new MutationObserver(patch).observe(document.body,{subtree:true,childList:true});
                          patch();
                        })();
                        </script>
                        """, height=0)

                        # ── Serialización robusta con sanitización ─────────────
                        def _sanitize_est_df(df):
                            """Uppercase + limpia espacios y chars especiales en texto."""
                            import re
                            df = df.copy()
                            if "codigo" in df.columns:
                                df["codigo"] = (df["codigo"].fillna("").astype(str)
                                                .str.strip().str.upper()
                                                .apply(lambda x: re.sub(r"[^A-Z0-9ÁÉÍÓÚÑ\-/]", "", x)))
                            if "obs" in df.columns:
                                df["obs"] = (df["obs"].fillna("").astype(str)
                                             .str.strip().str.upper())
                            if "crioscopia" in df.columns:
                                df["crioscopia"] = df["crioscopia"].fillna("").astype(str).str.strip()
                            return df

                        _clean_df = _sanitize_est_df(_edited_df_e)
                        _est_records = []
                        for _, _er in _clean_df.iterrows():
                            _rec = {}
                            for _ck in _clean_df.columns:
                                if _ck == "nombre_estacion":
                                    continue
                                _v = _er[_ck]
                                try:
                                    if pd.isna(_v):
                                        _rec[_ck] = None
                                    elif _ck == "volumen":
                                        _rec[_ck] = int(float(_v))
                                    elif hasattr(_v, "item"):
                                        _rec[_ck] = _v.item()
                                    else:
                                        _rec[_ck] = _v
                                except (TypeError, ValueError):
                                    _rec[_ck] = str(_v) if _v is not None else None
                            _est_records.append(_rec)
                        _edited_est_json = json.dumps(_est_records, ensure_ascii=False)

                        # ── Precalcular ponderados desde las estaciones editadas ──
                        _e_vols_p, _e_sum_st, _e_sum_ic = [], [], []
                        for _er2 in _est_records:
                            try:   _ev2 = float(_er2.get("volumen") or 0)
                            except: _ev2 = 0.0
                            try:   _es2 = float(str(_er2.get("solidos","") or "").replace(",","."))
                            except: _es2 = None
                            try:
                                _ec2_raw = str(_er2.get("crioscopia","") or "").strip()
                                _ec2 = float(_ec2_raw.replace(",",".")) if _ec2_raw else None
                            except Exception: _ec2 = None
                            if _ev2 > 0:
                                _e_vols_p.append(_ev2)
                                if _es2 is not None: _e_sum_st.append(_ev2 * _es2)
                                if _ec2 is not None: _e_sum_ic.append(_ev2 * _ec2)
                        _e_vol_total   = sum(_e_vols_p)
                        _e_vol_ests    = int(_e_vol_total) if _e_vol_total else 0
                        _e_diferencia  = int(_e_vol) - _e_vol_ests
                        _e_st_pond     = round(sum(_e_sum_st) / _e_vol_total, 2)  if _e_vol_total and _e_sum_st else ""
                        _e_ic_pond     = round(sum(_e_sum_ic) / _e_vol_total, 3)  if _e_vol_total and _e_sum_ic else ""

                        ec1, ec2, _ = st.columns([1.5, 1, 3])
                        with ec1:
                            if st.button("💾 GUARDAR CAMBIOS", type="primary",
                                         key="btn_save_edit", use_container_width=True):
                                try:
                                    _ests_cnt = len(json.loads(_edited_est_json) or [])
                                except Exception:
                                    _ests_cnt = 0
                                update_row_in_csv(_det_idx, {
                                    "fecha":             _e_fecha.strftime("%d/%m/%Y"),
                                    "ruta":              str(_e_ruta).upper(),
                                    "placa":             str(_e_placa).upper(),
                                    "conductor":         str(_e_cond).upper(),
                                    "volumen_declarado": int(_e_vol),
                                    "solidos_ruta":      round(float(_e_st), 2),
                                    "crioscopia_ruta":   round(float(_e_ic), 3),
                                    "vol_estaciones":    _e_vol_ests,
                                    "diferencia":        _e_diferencia,
                                    "st_pond":           _e_st_pond,
                                    "ic_pond":           _e_ic_pond,
                                    "estaciones_json":   _edited_est_json,
                                    "num_estaciones":    _ests_cnt,
                                })
                                st.session_state.pop(f"_h_est_cache_{st.session_state.get('admin_idx', 0)}", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with ec2:
                            if st.button("✖ CANCELAR", key="btn_cancel_edit",
                                         use_container_width=True):
                                st.session_state.pop(f"_h_est_cache_{st.session_state.get('admin_idx', 0)}", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                if not _edit_mode:
                    # ══ MODO VISTA: tarjetas KPI + tabla estática ════════

                    # ── Fila 1 de KPIs (tarjetas HTML) ───────────────────────
                    def _kpi_card(label, value, badge=None, badge_ok=True):  # noqa: E301
                        badge_html = ""
                        if badge:
                            bg = "#D4EDDA" if badge_ok else "#F8D7DA"
                            col = "#155724" if badge_ok else "#721C24"
                            badge_html = (f'<div style="margin-top:4px;font-size:.65rem;'
                                          f'font-weight:700;color:{col};background:{bg};'
                                          f'border-radius:4px;padding:1px 5px;display:inline-block;">'
                                          f'{badge}</div>')
                        return (
                            f'<div style="background:#fff;border:1px solid #dde6f0;border-radius:8px;'
                            f'padding:10px 12px;text-align:center;height:100%;">'
                            f'<div style="font-size:.62rem;font-weight:700;color:#6c8ca8;'
                            f'letter-spacing:.06em;margin-bottom:4px;">{label}</div>'
                            f'<div style="font-size:1.05rem;font-weight:800;color:#0056A3;">{value}</div>'
                            f'{badge_html}</div>'
                        )

                    try:    _v_vold = f"{int(float(_d_vold)):,} L"
                    except: _v_vold = "—"
                    try:
                        _st_v  = float(_d_st)
                        _st_ok = _st_v >= 12.60
                        _v_st  = f"{_st_v:.2f} %"
                        _b_st  = ("✔ CONFORME" if _st_ok else "✖ DESVIACIÓN", _st_ok)
                    except: _v_st = "—"; _b_st = (None, True)
                    try:
                        _ic_v  = float(_d_ic)
                        _ic_ok = -0.550 <= _ic_v <= -0.535
                        _v_ic  = f"{_ic_v:.3f} °C"
                        _b_ic  = ("✔ CONFORME" if _ic_ok else "✖ DESVIACIÓN", _ic_ok)
                    except: _v_ic = "—"; _b_ic = (None, True)
                    try:    _v_nest = str(int(float(_d_nest)))
                    except: _v_nest = str(_d_nest)

                    # Valores auxiliares para filas 1 y 2
                    try:    _v_vole = f"{int(float(_d_vole)):,} L"
                    except: _v_vole = "—"
                    try:
                        _dif_v  = int(float(_d_dif))
                        _dif_ok = abs(_dif_v) <= 20
                        _v_dif  = f"{_dif_v:+,} L"
                        _b_dif  = ("✔ OK" if _dif_ok else "⚠ DIFERENCIA", _dif_ok)
                    except: _v_dif = "—"; _b_dif = (None, True)
                    try:
                        _stpv  = float(_d_stpond)
                        _v_stp = f"{_stpv:.2f} %"
                    except: _v_stp = "—"
                    try:
                        _dif_st_v = float(_d_st) - float(_d_stpond)
                        _v_dif_st = f"{_dif_st_v:+.2f} %"
                    except: _v_dif_st = "—"

                    # IC PONDERADO solo si TODAS las estaciones tienen crioscopia válida
                    _todas_con_ic = False
                    try:
                        _raw_check = str(_drow.get("estaciones_json", "") or "").strip()
                        if _raw_check:
                            _ests_check = json.loads(_raw_check)
                            if _ests_check:
                                _todas_con_ic = all(
                                    (lambda v: v is not None)(
                                        (lambda s: float(s.replace(",", "."))
                                         if s and str(s).strip() not in ("", "None", "nan") else None
                                        )(str(_ec.get("crioscopia", "") or ""))
                                    )
                                    for _ec in _ests_check
                                )
                    except Exception:
                        _todas_con_ic = False

                    # ── Filas KPI según tipo de registro ─────────────────────
                    if _tipo_reg == "TRANSUIZA":
                        # ── Vista TRANSUIZA ────────────────────────────────────
                        try:
                            _v_stcar = f"{float(str(_d_st_car or 0).replace(',','.')):.2f} %"
                        except: _v_stcar = "—"
                        try:
                            _stm_val = float(str(_d_st or 0).replace(",","."))
                            _stm_ok  = _stm_val >= 12.60
                            _v_stm   = f"{_stm_val:.2f} %"
                            _b_stm   = ("✔ CONFORME" if _stm_ok else "✖ DESVIACIÓN", _stm_ok)
                        except: _v_stm = "—"; _b_stm = (None, True)
                        try:
                            _v_grasa_t = f"{float(str(_d_grasa or 0).replace(',','.')):.2f} %"
                        except: _v_grasa_t = "—"
                        try:
                            _v_prot_t = f"{float(str(_d_prot or 0).replace(',','.')):.2f} %"
                        except: _v_prot_t = "—"
                        try:
                            _dif_s_v  = float(str(_d_dif_s or 0).replace(",","."))
                            _dif_s_ok = abs(_dif_s_v) <= 0.5
                            _v_dif_s  = f"{_dif_s_v:+.2f} %"
                            _b_dif_s  = ("✔ OK" if _dif_s_ok else "⚠ DIFERENCIA", _dif_s_ok)
                        except: _v_dif_s = "—"; _b_dif_s = (None, True)

                        _trans_kpi_html = (
                            '<div style="display:grid;grid-template-columns:repeat(5,1fr);'
                            'gap:8px;margin-bottom:14px;">'
                            + _kpi_card("PLACA",              _d_placa or "—")
                            + _kpi_card("ST CARROTANQUE",     _v_stcar)
                            + _kpi_card("ST MUESTRA",         _v_stm, _b_stm[0], _b_stm[1])
                            + _kpi_card("GRASA",              _v_grasa_t)
                            + _kpi_card("PROTEÍNA",           _v_prot_t)
                            + '</div>'
                            + '<div style="display:grid;grid-template-columns:repeat(1,1fr);'
                            'gap:8px;margin-bottom:14px;">'
                            + _kpi_card("DIF. DE SÓLIDOS (CARROTANQUE − MUESTRA)", _v_dif_s, _b_dif_s[0], _b_dif_s[1])
                            + '</div>'
                        )
                        st.markdown(_trans_kpi_html, unsafe_allow_html=True)

                    else:
                        # ── Fila 1: PLACA · CONDUCTOR · VOL. DECL. · ST RUTA · ST POND · ΔST
                        _row1_html = (
                            '<div style="display:grid;grid-template-columns:repeat(6,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_card("PLACA",          _d_placa or "—")
                            + _kpi_card("CONDUCTOR",      (_d_cond[:16]+"…") if len(_d_cond)>16 else (_d_cond or "—"))
                            + _kpi_card("VOL. DECLARADO", _v_vold)
                            + _kpi_card("ST RUTA",        _v_st,  _b_st[0], _b_st[1])
                            + _kpi_card("ST PONDERADO",   _v_stp)
                            + _kpi_card("ΔST (RUTA−POND)",_v_dif_st)
                            + '</div>'
                        )

                    # ── Fila 2: Nº EST · DIF. VOL. · VOL. EST · IC RUTA · [IC POND · ΔIC]
                        _row2_cards = [
                            _kpi_card("Nº ESTACIONES",  _v_nest),
                            _kpi_card("DIF. VOLUMEN",   _v_dif,  _b_dif[0], _b_dif[1]),
                            _kpi_card("VOL. ESTACIONES",_v_vole),
                            _kpi_card("IC RUTA",        _v_ic,   _b_ic[0],  _b_ic[1]),
                        ]
                        if _todas_con_ic:
                            try:
                                _icpv     = float(_d_icpond)
                                _dif_ic_v = float(_d_ic) - _icpv
                                _row2_cards.append(_kpi_card("IC PONDERADO",   f"{_icpv:.3f} °C"))
                                _row2_cards.append(_kpi_card("ΔIC (RUTA−POND)",f"{_dif_ic_v:+.3f} °C"))
                            except Exception:
                                pass

                        _ncols2 = len(_row2_cards)
                        _row2_html = (
                            f'<div style="display:grid;grid-template-columns:repeat({_ncols2},1fr);'
                            f'gap:8px;margin-bottom:14px;">'
                            + "".join(_row2_cards)
                            + '</div>'
                        )

                        st.markdown(_row1_html + _row2_html, unsafe_allow_html=True)

                        # ── Tabla de estaciones ───────────────────────────────────
                        _est_json_raw = str(_drow.get("estaciones_json", "") or "").strip()
                        if _est_json_raw:
                            try:
                                _ests = json.loads(_est_json_raw)
                                if _ests:
                                    st.markdown(
                                        "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                        "letter-spacing:.05em;margin-bottom:4px;'>"
                                        "📋 CALIDAD POR ESTACIÓN</div>",
                                        unsafe_allow_html=True,
                                    )

                                    def _pn(x):
                                        try: return float(str(x).replace(",", "."))
                                        except: return None

                                    _cat_det = load_catalogo()
                                    _cat_det_map = dict(zip(_cat_det["codigo"], _cat_det["nombre"]))
                                    _det_rows = []
                                    for _e in _ests:
                                        _ev  = _pn(_e.get("volumen"))
                                        _est_v = _pn(_e.get("solidos"))
                                        _eic = _pn(_e.get("crioscopia"))
                                        _pst = round(_ev * _est_v, 2) if _ev is not None and _est_v is not None else None
                                        _pic = round(_ev * _eic,   3) if _ev is not None and _eic is not None else None
                                        _cod_e = _e.get("codigo", "") or ""
                                        _det_rows.append({
                                            "CÓDIGO":           _cod_e,
                                            "GRASA (%)":        _pn(_e.get("grasa")),
                                            "SÓLIDOS TOT. (%)": _est_v,
                                            "PROTEÍNA (%)":     _pn(_e.get("proteina")),
                                            "CRIOSCOPIA":       _eic,
                                            "AGUA (%)":         _pn(_e.get("agua_pct")),
                                            "VOLUMEN (L)":      int(_ev) if _ev is not None else None,
                                            "POND ST":          _pst,
                                            "IC POND":          _pic,
                                            "ALC.":             _e.get("alcohol",        "N/A") or "N/A",
                                            "CLOR.":            _e.get("cloruros",       "N/A") or "N/A",
                                            "NEUT.":            _e.get("neutralizantes", "N/A") or "N/A",
                                            "OBSERVACIONES":    _e.get("obs", "") or "",
                                            "NOMBRE ESTACIÓN":  _cat_det_map.get(_cod_e.strip(), ""),
                                        })

                                    _df_det = pd.DataFrame(_det_rows)

                                    _RED_EST = "background-color:#FFC7CE;color:#9C0006;font-weight:700"

                                    def _color_est(row):
                                        styles = [""] * len(row)
                                        cols = list(row.index)
                                        try:
                                            if row.get("SÓLIDOS TOT. (%)") is not None and 0 < float(row["SÓLIDOS TOT. (%)"]) < 12.60:
                                                if "SÓLIDOS TOT. (%)" in cols:
                                                    styles[cols.index("SÓLIDOS TOT. (%)")] = _RED_EST
                                        except Exception: pass
                                        try:
                                            _icv = row.get("CRIOSCOPIA")
                                            if _icv is not None:
                                                _icf = float(_icv)
                                                if _icf > -0.530:
                                                    if "CRIOSCOPIA" in cols:
                                                        styles[cols.index("CRIOSCOPIA")] = _RED_EST
                                                    if "AGUA (%)" in cols:
                                                        styles[cols.index("AGUA (%)")] = _RED_EST
                                                elif _icf < -0.550:
                                                    if "CRIOSCOPIA" in cols:
                                                        styles[cols.index("CRIOSCOPIA")] = _RED_EST
                                        except Exception: pass
                                        for _qcol in ("ALC.", "CLOR.", "NEUT."):
                                            try:
                                                if row.get(_qcol) == "+" and _qcol in cols:
                                                    styles[cols.index(_qcol)] = _RED_EST
                                            except Exception: pass
                                        return styles

                                    _fmt_det = {
                                        "GRASA (%)":        "{:.2f}",
                                        "SÓLIDOS TOT. (%)": "{:.2f}",
                                        "PROTEÍNA (%)":     "{:.2f}",
                                        "CRIOSCOPIA":       "{:.3f}",
                                        "AGUA (%)":         "{:.1f}",
                                        "POND ST":          "{:.2f}",
                                        "IC POND":          "{:.3f}",
                                    }
                                    st.dataframe(
                                        _df_det.style
                                               .apply(_color_est, axis=1)
                                               .format(_fmt_det, na_rep="—"),
                                        use_container_width=True,
                                        hide_index=True,
                                        height=min(38 + 35 * len(_det_rows), 420),
                                        column_config={
                                            "CÓDIGO":           st.column_config.TextColumn("CÓDIGO",          width="small"),
                                            "GRASA (%)":        st.column_config.NumberColumn("GRASA (%)",      width="small",  format="%.2f"),
                                            "SÓLIDOS TOT. (%)": st.column_config.NumberColumn("ST (%)",         width="small",  format="%.2f"),
                                            "PROTEÍNA (%)":     st.column_config.NumberColumn("PROT. (%)",      width="small",  format="%.2f"),
                                            "CRIOSCOPIA":       st.column_config.NumberColumn("CRIOS.",         width="small",  format="%.3f"),
                                            "AGUA (%)":         st.column_config.NumberColumn("AGUA (%)",       width="small",  format="%.1f"),
                                            "VOLUMEN (L)":      st.column_config.NumberColumn("VOL. (L)",       width="small",  format="%d"),
                                            "POND ST":          st.column_config.NumberColumn("POND ST",        width="small",  format="%.2f"),
                                            "IC POND":          st.column_config.NumberColumn("IC POND",        width="small",  format="%.3f"),
                                            "ALC.":             st.column_config.TextColumn("ALC.",             width="small"),
                                            "CLOR.":            st.column_config.TextColumn("CLOR.",            width="small"),
                                            "NEUT.":            st.column_config.TextColumn("NEUT.",            width="small"),
                                            "OBSERVACIONES":    st.column_config.TextColumn("OBSERVACIONES",    width="medium"),
                                        },
                                    )
                            except Exception:
                                pass
                        else:
                            st.caption("Esta ruta no tiene datos de estaciones registrados.")

                    # ── Galería de fotos almacenadas ──────────────────────────
                    _fotos_raw = str(_drow.get("fotos_json", "") or "").strip()
                    if _fotos_raw and _fotos_raw not in ("[]", ""):
                        try:
                            _fotos_list = json.loads(_fotos_raw)
                        except Exception:
                            _fotos_list = []
                        _fotos_existentes = [p for p in _fotos_list if os.path.exists(p)]
                        if _fotos_existentes:
                            st.markdown(
                                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                "letter-spacing:.05em;margin:10px 0 6px;'>"
                                "📷 IMÁGENES DE MUESTRAS</div>",
                                unsafe_allow_html=True,
                            )
                            _cols_fotos = st.columns(min(len(_fotos_existentes), 4))
                            for _fi, _fp in enumerate(_fotos_existentes):
                                with _cols_fotos[_fi % 4]:
                                    st.image(_fp, use_container_width=True)

            # ── Acción ELIMINAR ────────────────────────────────────────────
            accion_activa = st.session_state.get("admin_accion")
            idx_activo    = st.session_state.get("admin_idx")

            _tiene_indices = (idx_activo is not None or
                              bool(st.session_state.get("admin_idxs")))
            if accion_activa == "eliminar" and _tiene_indices:
                with st.container(border=True):
                    # ── Confirmación de borrado (1 o varios) ─────
                    _idxs_del = st.session_state.get("admin_idxs") or (
                        [idx_activo] if idx_activo is not None else []
                    )
                    n_del = len(_idxs_del)
                    st.markdown(
                        f"<div style='font-weight:700;color:#9C0006;margin-bottom:4px;'>"
                        f"🗑️ ¿Confirmar eliminación de "
                        f"{'1 registro' if n_del == 1 else f'{n_del} registros'}?</div>",
                        unsafe_allow_html=True,
                    )
                    if n_del == 1 and idx_activo is not None:
                        row_a = df_hist.loc[idx_activo] if idx_activo in df_hist.index else {}
                        st.markdown(
                            f"**Fecha:** {row_a.get('fecha','')} &nbsp;·&nbsp; "
                            f"**Ruta:** {row_a.get('ruta','')} &nbsp;·&nbsp; "
                            f"**Placa:** {row_a.get('placa','')}",
                        )
                    else:
                        for _di in _idxs_del:
                            if _di in df_hist.index:
                                _r = df_hist.loc[_di]
                                st.markdown(
                                    f"· **{_r.get('fecha','')}** — "
                                    f"{_r.get('ruta','')} / {_r.get('placa','')}",
                                )
                    dc1, dc2, _ = st.columns([1.5, 1, 3])
                    with dc1:
                        if st.button("🗑️ CONFIRMAR", type="primary",
                                     key="btn_confirm_del", use_container_width=True):
                            if filtro_tipo == "SEGUIMIENTOS":
                                delete_seg_rows(_idxs_del)
                            else:
                                delete_rows_from_csv(_idxs_del)
                            st.session_state.admin_accion = None
                            st.session_state.admin_idx    = None
                            st.session_state.admin_idxs   = []
                            st.rerun()
                    with dc2:
                        if st.button("✖ CANCELAR", key="btn_cancel_del",
                                     use_container_width=True):
                            st.session_state.admin_accion = None
                            st.session_state.admin_idx    = None
                            st.session_state.admin_idxs   = []
                            st.rerun()



elif st.session_state.pagina_activa == "DASHBOARD":
    # ── DASHBOARD DE CALIDAD ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        """<div style="font-size:1rem;font-weight:700;color:#0056A3;
           letter-spacing:.04em;margin:6px 0 10px 0;">📊 DASHBOARD DE CALIDAD</div>""",
        unsafe_allow_html=True,
    )

    _ST_MIN = 12.60
    _IC_MIN = -0.550
    _IC_MAX = -0.535

    _df_raw  = load_historial()
    _df_dash = _df_raw[_df_raw["tipo_seguimiento"].isin(["RUTAS", "TRANSUIZA"])].copy()

    if _df_dash.empty:
        st.info("Aún no hay registros de RUTAS o TRANSUIZA para mostrar en el dashboard.")
    else:
        _df_dash["_fecha_dt"] = pd.to_datetime(
            _df_dash["fecha"], format="%d/%m/%Y", errors="coerce"
        )
        _df_dash = _df_dash.dropna(subset=["_fecha_dt"])
        _df_dash["solidos_ruta"]      = pd.to_numeric(_df_dash["solidos_ruta"],      errors="coerce")
        _df_dash["crioscopia_ruta"]   = pd.to_numeric(_df_dash["crioscopia_ruta"],   errors="coerce")
        _df_dash["volumen_declarado"] = pd.to_numeric(_df_dash["volumen_declarado"], errors="coerce")

        _f_min = _df_dash["_fecha_dt"].min().date()
        _f_max = _df_dash["_fecha_dt"].max().date()

        _fc1, _fc2, _fc3 = st.columns([2, 2, 2])
        with _fc1:
            _dash_desde = st.date_input(
                "DESDE", value=_f_min, min_value=_f_min, max_value=_f_max,
                key="dash_desde", format="DD/MM/YYYY",
            )
        with _fc2:
            _dash_hasta = st.date_input(
                "HASTA", value=_f_max, min_value=_f_min, max_value=_f_max,
                key="dash_hasta", format="DD/MM/YYYY",
            )
        with _fc3:
            _dash_tipo = st.selectbox("TIPO", ["TODOS", "RUTAS", "TRANSUIZA"], key="dash_tipo")

        _df_f = _df_dash[
            (_df_dash["_fecha_dt"].dt.date >= _dash_desde) &
            (_df_dash["_fecha_dt"].dt.date <= _dash_hasta)
        ].copy()
        if _dash_tipo != "TODOS":
            _df_f = _df_f[_df_f["tipo_seguimiento"] == _dash_tipo]

        if _df_f.empty:
            st.warning("No hay datos para el período y tipo seleccionados.")
        else:
            _df_f["_st_ok"] = _df_f["solidos_ruta"]  >= _ST_MIN
            _df_f["_ic_ok"] = (
                (_df_f["crioscopia_ruta"] >= _IC_MIN) &
                (_df_f["crioscopia_ruta"] <= _IC_MAX)
            )
            _df_f["_conf"]   = _df_f["_st_ok"] & _df_f["_ic_ok"]
            _df_f["_estado"] = _df_f["_conf"].map({True: "CONFORME", False: "NO CONFORME"})

            _n_tot   = len(_df_f)
            _n_conf  = int(_df_f["_conf"].sum())
            _pct     = (_n_conf / _n_tot * 100) if _n_tot else 0
            _avg_st  = float(_df_f["solidos_ruta"].mean())
            _avg_ic  = float(_df_f["crioscopia_ruta"].mean())
            _vol_tot = int(_df_f["volumen_declarado"].dropna().sum())

            # ── KPIs ──────────────────────────────────────────────────────
            _k1, _k2, _k3, _k4 = st.columns(4)
            _k1.metric("CUMPLIMIENTO",  f"{_pct:.1f}%",
                       f"{_n_conf} de {_n_tot} registros", delta_color="off")
            _k2.metric("PROMEDIO ST",   f"{_avg_st:.2f}%",
                       f"{'✔ OK' if _avg_st >= _ST_MIN else '✖ BAJO'} — mín {_ST_MIN}%",
                       delta_color="off")
            _k3.metric("PROMEDIO IC",   f"{_avg_ic:.3f} °C",
                       f"Rango [{_IC_MIN}, {_IC_MAX}]", delta_color="off")
            _k4.metric("VOLUMEN TOTAL", f"{_vol_tot:,} L",
                       f"{_n_tot} registros en período", delta_color="off")

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

            # ── Barras por ruta ────────────────────────────────────────────
            _df_agg = (
                _df_f.groupby("ruta")
                .agg(
                    st_avg   =("solidos_ruta",    "mean"),
                    ic_avg   =("crioscopia_ruta",  "mean"),
                    n        =("solidos_ruta",    "count"),
                    pct_conf =("_conf",            "mean"),
                )
                .reset_index()
            )
            _df_agg["estado"] = _df_agg["pct_conf"].apply(
                lambda x: "CONFORME" if x >= 0.5 else "NO CONFORME"
            )

            _bc1, _bc2 = st.columns(2)

            with _bc1:
                st.markdown(
                    "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                    "margin-bottom:2px;'>SÓLIDOS TOTALES POR RUTA (%)</div>",
                    unsafe_allow_html=True,
                )
                _bar_st = (
                    alt.Chart(_df_agg)
                    .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                    .encode(
                        x=alt.X("ruta:N", title=None,
                                sort=alt.EncodingSortField("st_avg", order="descending"),
                                axis=alt.Axis(labelAngle=-30, labelFontSize=10)),
                        y=alt.Y("st_avg:Q", title="ST Prom. (%)",
                                scale=alt.Scale(domain=[12.0, 13.5])),
                        color=alt.Color(
                            "estado:N",
                            scale=alt.Scale(
                                domain=["CONFORME", "NO CONFORME"],
                                range=["#0056A3", "#EF4444"],
                            ),
                            legend=alt.Legend(title="Estado"),
                        ),
                        tooltip=[
                            alt.Tooltip("ruta:N",   title="Ruta"),
                            alt.Tooltip("st_avg:Q", title="ST Prom.", format=".2f"),
                            alt.Tooltip("n:Q",      title="Registros"),
                        ],
                    )
                )
                _rule_st = (
                    alt.Chart(pd.DataFrame({"y": [_ST_MIN]}))
                    .mark_rule(color="#EF4444", strokeDash=[5, 3], strokeWidth=1.5)
                    .encode(y="y:Q")
                )
                st.altair_chart(
                    (_bar_st + _rule_st).properties(height=240).configure_view(strokeWidth=0),
                    use_container_width=True,
                )

            with _bc2:
                st.markdown(
                    "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                    "margin-bottom:2px;'>ÍNDICE CRIOSCÓPICO POR RUTA (°C)</div>",
                    unsafe_allow_html=True,
                )
                _bar_ic = (
                    alt.Chart(_df_agg)
                    .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                    .encode(
                        x=alt.X("ruta:N", title=None,
                                sort=alt.EncodingSortField("ic_avg", order="ascending"),
                                axis=alt.Axis(labelAngle=-30, labelFontSize=10)),
                        y=alt.Y("ic_avg:Q", title="IC Prom. (°C)",
                                scale=alt.Scale(domain=[-0.570, -0.510])),
                        color=alt.Color(
                            "estado:N",
                            scale=alt.Scale(
                                domain=["CONFORME", "NO CONFORME"],
                                range=["#10B981", "#EF4444"],
                            ),
                            legend=alt.Legend(title="Estado"),
                        ),
                        tooltip=[
                            alt.Tooltip("ruta:N",   title="Ruta"),
                            alt.Tooltip("ic_avg:Q", title="IC Prom.", format=".3f"),
                            alt.Tooltip("n:Q",      title="Registros"),
                        ],
                    )
                )
                _ref_icmin = (
                    alt.Chart(pd.DataFrame({"y": [_IC_MIN]}))
                    .mark_rule(color="#EF4444", strokeDash=[5, 3], strokeWidth=1.5)
                    .encode(y="y:Q")
                )
                _ref_icmax = (
                    alt.Chart(pd.DataFrame({"y": [_IC_MAX]}))
                    .mark_rule(color="#F97316", strokeDash=[5, 3], strokeWidth=1.5)
                    .encode(y="y:Q")
                )
                st.altair_chart(
                    (_bar_ic + _ref_icmin + _ref_icmax)
                    .properties(height=240)
                    .configure_view(strokeWidth=0),
                    use_container_width=True,
                )

            # ── Tendencia diaria ST ───────────────────────────────────────
            st.markdown(
                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                "margin:4px 0 2px;'>TENDENCIA DIARIA — SÓLIDOS TOTALES (%)</div>",
                unsafe_allow_html=True,
            )
            _df_daily = (
                _df_f.groupby(_df_f["_fecha_dt"].dt.date)
                .agg(st_avg=("solidos_ruta", "mean"), n=("solidos_ruta", "count"))
                .reset_index()
                .rename(columns={"_fecha_dt": "fecha"})
            )
            _df_daily["fecha"] = pd.to_datetime(_df_daily["fecha"])

            _line_st = (
                alt.Chart(_df_daily)
                .mark_line(point=True, color="#0056A3", strokeWidth=2)
                .encode(
                    x=alt.X("fecha:T", title="Fecha"),
                    y=alt.Y("st_avg:Q", title="ST Prom. (%)",
                            scale=alt.Scale(domain=[12.0, 13.5])),
                    tooltip=[
                        alt.Tooltip("fecha:T",  title="Fecha",    format="%d/%m/%Y"),
                        alt.Tooltip("st_avg:Q", title="ST Prom.", format=".2f"),
                        alt.Tooltip("n:Q",      title="Registros"),
                    ],
                )
            )
            _rule_st2 = (
                alt.Chart(pd.DataFrame({"y": [_ST_MIN]}))
                .mark_rule(color="#EF4444", strokeDash=[5, 3], strokeWidth=1.5)
                .encode(y="y:Q")
            )
            st.altair_chart(
                (_line_st + _rule_st2)
                .properties(height=190)
                .configure_view(strokeWidth=0),
                use_container_width=True,
            )

            # ── Últimos 10 registros ──────────────────────────────────────
            st.markdown(
                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                "margin:4px 0 2px;'>ÚLTIMOS 10 REGISTROS</div>",
                unsafe_allow_html=True,
            )
            _df_rec = (
                _df_f.sort_values("_fecha_dt", ascending=False)
                .head(10)[["fecha", "ruta", "placa", "solidos_ruta",
                           "crioscopia_ruta", "volumen_declarado", "_estado"]]
                .rename(columns={
                    "fecha":             "FECHA",
                    "ruta":              "RUTA",
                    "placa":             "PLACA",
                    "solidos_ruta":      "ST (%)",
                    "crioscopia_ruta":   "IC (°C)",
                    "volumen_declarado": "VOLUMEN (L)",
                    "_estado":           "ESTADO",
                })
                .reset_index(drop=True)
            )

            def _color_estado(row):
                if row["ESTADO"] == "NO CONFORME":
                    return ["background-color:#FEE2E2;color:#B91C1C"] * len(row)
                if row["ESTADO"] == "CONFORME":
                    return ["background-color:#DCFCE7;color:#15803D"] * len(row)
                return [""] * len(row)

            st.dataframe(
                _df_rec.style.apply(_color_estado, axis=1),
                use_container_width=True,
                hide_index=True,
            )


    save_draft_state()
